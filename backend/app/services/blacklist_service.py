import csv
import hashlib
from datetime import datetime
from io import BytesIO, StringIO
from typing import Iterable, Optional

import xlrd
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.blacklist import BlacklistEntry
from app.models.user import User


def normalize_identity_value(value: Optional[str]) -> str:
    return str(value or "").strip()


def is_md5_text(value: Optional[str]) -> bool:
    text = normalize_identity_value(value).lower()
    return len(text) == 32 and all(char in "0123456789abcdef" for char in text)


def md5_text(value: Optional[str]) -> str:
    text = normalize_identity_value(value)
    if not text:
        return ""
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def build_match_values(value: Optional[str]) -> tuple[str, str]:
    text = normalize_identity_value(value)
    if not text:
        return "", ""
    return ("", text.lower()) if is_md5_text(text) else (text, md5_text(text))


async def find_blacklist_hit(
    db: AsyncSession,
    *,
    phone: Optional[str] = None,
    id_card_num: Optional[str] = None,
) -> Optional[BlacklistEntry]:
    phone_plain, phone_md5 = build_match_values(phone)
    id_plain, id_md5 = build_match_values(id_card_num)
    filters = [BlacklistEntry.removed_at.is_(None)]
    match_filters = []
    if phone_plain:
        match_filters.append(BlacklistEntry.phone == phone_plain)
    if phone_md5:
        match_filters.append(BlacklistEntry.phone_md5 == phone_md5)
        match_filters.append(BlacklistEntry.phone == phone_md5)
    if id_plain:
        match_filters.append(BlacklistEntry.id_card_num == id_plain)
    if id_md5:
        match_filters.append(BlacklistEntry.id_card_md5 == id_md5)
        match_filters.append(BlacklistEntry.id_card_num == id_md5)
    if not match_filters:
        return None
    stmt = select(BlacklistEntry).where(*filters, or_(*match_filters)).order_by(BlacklistEntry.created_at.desc())
    return (await db.execute(stmt)).scalars().first()


async def refresh_user_blacklist_status(db: AsyncSession, user: User) -> Optional[BlacklistEntry]:
    hit = await find_blacklist_hit(db, phone=user.phone, id_card_num=user.id_card_num)
    user.blacklist_checked_at = datetime.now()
    user.blacklist_hit = bool(hit)
    user.blacklist_reason = hit.reason or hit.source if hit else None
    if hit:
        user.approved_limit = 0
        user.available_credit_limit = 0
    return hit


async def add_blacklist_entry(
    db: AsyncSession,
    *,
    name: Optional[str] = None,
    phone: Optional[str] = None,
    id_card_num: Optional[str] = None,
    source: str = "MANUAL",
    reason: Optional[str] = None,
    created_by: Optional[str] = None,
) -> BlacklistEntry:
    phone_text = normalize_identity_value(phone)
    id_text = normalize_identity_value(id_card_num)
    if not phone_text and not id_text:
        raise HTTPException(status_code=400, detail="手机号和身份证号至少填写一个")
    phone_plain, phone_hash = build_match_values(phone_text)
    id_plain, id_hash = build_match_values(id_text)
    existed = await find_blacklist_hit(db, phone=phone_text, id_card_num=id_text)
    if existed:
        return existed
    now = datetime.now()
    entry = BlacklistEntry(
        name=normalize_identity_value(name) or None,
        phone=phone_plain or (phone_text.lower() if is_md5_text(phone_text) else None),
        id_card_num=id_plain or (id_text.lower() if is_md5_text(id_text) else None),
        phone_md5=phone_hash or (phone_text.lower() if is_md5_text(phone_text) else None),
        id_card_md5=id_hash or (id_text.lower() if is_md5_text(id_text) else None),
        source=source,
        reason=reason,
        created_by=created_by,
        created_at=now,
    )
    db.add(entry)
    await db.flush()
    return entry


async def blacklist_user(
    db: AsyncSession,
    user: User,
    *,
    source: str,
    reason: Optional[str],
    created_by: Optional[str],
) -> BlacklistEntry:
    entry = await add_blacklist_entry(
        db,
        name=user.name,
        phone=user.phone,
        id_card_num=user.id_card_num,
        source=source,
        reason=reason,
        created_by=created_by,
    )
    user.blacklist_hit = True
    user.blacklist_reason = reason or source
    user.blacklist_checked_at = datetime.now()
    user.approved_limit = 0
    user.available_credit_limit = 0
    return entry


async def remove_user_from_blacklist(
    db: AsyncSession,
    user: User,
    *,
    removed_by: Optional[str],
    reason: Optional[str] = None,
) -> int:
    phone_plain, phone_hash = build_match_values(user.phone)
    id_plain, id_hash = build_match_values(user.id_card_num)
    match_filters = []
    for column, value in (
        (BlacklistEntry.phone, phone_plain),
        (BlacklistEntry.phone, phone_hash),
        (BlacklistEntry.phone_md5, phone_hash),
        (BlacklistEntry.id_card_num, id_plain),
        (BlacklistEntry.id_card_num, id_hash),
        (BlacklistEntry.id_card_md5, id_hash),
    ):
        if value:
            match_filters.append(column == value)
    if not match_filters:
        return 0
    entries = (
        await db.execute(select(BlacklistEntry).where(BlacklistEntry.removed_at.is_(None), or_(*match_filters)))
    ).scalars().all()
    now = datetime.now()
    for entry in entries:
        entry.removed_at = now
        entry.removed_by = removed_by
        entry.remove_reason = reason
    user.blacklist_hit = False
    user.blacklist_reason = None
    user.blacklist_checked_at = now
    return len(entries)


def serialize_blacklist_entry(entry: BlacklistEntry) -> dict:
    return {
        "id": entry.id,
        "name": entry.name,
        "phone": entry.phone,
        "id_card_num": entry.id_card_num,
        "source": entry.source,
        "reason": entry.reason,
        "created_by": entry.created_by,
        "created_at": entry.created_at,
    }


def _sheet_rows_from_upload(file: UploadFile) -> Iterable[list[str]]:
    content = file.file.read()
    file.file.close()
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            yield ["" if value is None else str(value).strip() for value in row]
        return
    if filename.endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        for row_idx in range(1, sheet.nrows):
            yield [str(sheet.cell_value(row_idx, col_idx)).strip() for col_idx in range(min(3, sheet.ncols))]
        return
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.reader(StringIO(text))
        next(reader, None)
        yield from ([str(value).strip() for value in row[:3]] for row in reader)
        return
    if filename.endswith(".txt"):
        text = content.decode("utf-8-sig")
        for index, line in enumerate(text.splitlines()):
            if index == 0 and any(title in line for title in ("姓名", "手机号", "身份证")):
                continue
            if not line.strip():
                continue
            delimiter = "," if "," in line else "\t"
            yield [part.strip() for part in line.split(delimiter)[:3]]
        return
    raise HTTPException(status_code=400, detail="仅支持 xlsx、xls、txt、csv 文件")


async def upload_blacklist_entries(db: AsyncSession, file: UploadFile, *, created_by: Optional[str]) -> dict:
    rows = list(_sheet_rows_from_upload(file))
    if not rows:
        raise HTTPException(status_code=400, detail="上传文件内容不能为空")
    created = 0
    skipped = 0
    errors = []
    for index, row in enumerate(rows, start=2):
        name = row[0] if len(row) > 0 else ""
        phone = row[1] if len(row) > 1 else ""
        id_card = row[2] if len(row) > 2 else ""
        if not phone and not id_card:
            errors.append({"row": index, "reason": "手机号和身份证号至少填写一个"})
            continue
        before = await find_blacklist_hit(db, phone=phone, id_card_num=id_card)
        await add_blacklist_entry(
            db,
            name=name,
            phone=phone,
            id_card_num=id_card,
            source="UPLOAD",
            reason="批量导入",
            created_by=created_by,
        )
        if before:
            skipped += 1
        else:
            created += 1
    return {"created": created, "skipped": skipped, "errors": errors}
