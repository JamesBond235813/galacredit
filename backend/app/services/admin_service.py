import asyncio
import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Optional
import string
import secrets

import xlrd
from fastapi import Depends, File, HTTPException, Query, UploadFile, WebSocket, WebSocketDisconnect, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import case, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.api.deps import get_admin_by_token_async, get_current_admin_async
from app.core.config import settings
from app.core.database import get_async_db
from app.core.security import create_access_token, get_password_hash, verify_password
from app.core.trace import new_trace_id, reset_trace_id, set_trace_id
from app.models.admin import Admin
from app.models.blacklist import BlacklistEntry
from app.models.channel import Channel
from app.models.ecard_pool import EcardPool
from app.models.compliance_rule import ComplianceRule
from app.models.loan import Loan
from app.models.momo_transaction import MomoTransaction
from app.models.loan_ecard import LoanEcard
from app.models.loan_transaction import LoanTransaction
from app.models.overdue_fee_config import OverdueFeeConfig
from app.models.product import Product
from app.models.risk_composite_report import RiskCompositeReport
from app.models.user import User
from app.models.user_event import UserEvent
from app.models.ops_history import ConfigChangeHistory, MessageTemplate
from app.schemas.admin import (
    AdminLogin,
    AdminTokenResponse,
    RegisterUserRequest,
    ResetUserPasswordRequest,
    AdminResponse,
    AdminUserCreateRequest,
    AdminUserItemResponse,
    AdminUserUpdateRequest,
    PaginatedAdminUserResponse,
    AdminChangePasswordRequest,
)
from app.schemas.channel import (
    BusinessAdvisorItemResponse,
    ChannelCreateRequest,
    ChannelUpdateRequest,
    PaginatedChannelResponse,
)
from app.schemas.loan import (
    AdminStatsResponse,
    ApprovedCreditSetRequest,
    AvailableCreditAdjustRequest,
    DisburseRequest,
    EcardPoolCreateRequest,
    EcardPoolUpdateRequest,
    LoanAssigneeItemResponse,
    LoanAssignmentResponse,
    LoanAssignRequest,
    LoanLedgerResponse,
    LoanFinanceReconcileRequest,
    LoanFollowUpRequest,
    ProjectCashInsightResponse,
    ProductCreateRequest,
    ProductUpdateRequest,
    RepayAttemptAckResponse,
    RepaymentStatsResponse,
    LoanReviewRequest,
    LoanUpdateRequest,
    OverdueDisplayRequest,
    OverdueFeeConfigCreateRequest,
    PaginatedEcardPoolResponse,
    PaginatedProductResponse,
    PaginatedLoanResponse,
)
from app.schemas.risk import AdminRiskReportRequest, AdminRiskSingleReportRequest, RiskReportResponse
from app.schemas.user import PaginatedUserResponse, UserDetailResponse
from app.services.audit import log_user_event_async
from app.services.admin_permissions import (
    ALL_ADMIN_PERMISSION_KEYS,
    admin_has_permission,
    parse_admin_roles,
    parse_admin_permissions,
    resolve_admin_permissions,
    resolve_permissions_from_roles,
    serialize_admin_permissions,
    serialize_admin_roles,
)
from app.services.channel_service import (
    build_daily_channel_invite_code,
    build_channel_metrics,
    build_channel_summary,
    get_channel_by_name_async,
    normalize_channel_name,
    normalize_channel_disbursement_mode,
    normalize_channel_review_mode,
    normalize_channel_status,
    serialize_channel_landing,
)
from app.services.loan_amounts import (
    DEFAULT_FEE_RATE,
    calculate_remaining_repayment_amount,
    calculate_total_repayment_amount,
    normalize_fee_rate,
    normalize_term_days,
    serialize_loan_snapshot,
    sync_loan_fee_fields,
    normalize_installment_ratios,
)
from app.services.loan_ledger import (
    create_disbursement_transaction_async,
    ensure_installment_records_async,
    get_loan_ledger_snapshot,
    register_other_fee_async,
    register_reduction_async,
    register_repayment_async,
    serialize_transaction,
    sync_loan_repayment_state,
)
from app.services.loan_flow import (
    get_latest_loan_async,
    get_latest_normal_settled_loan,
    get_relend_count,
    get_relend_label,
)
from app.services.loan_assignment import (
    COLLECTION_TRANSFER_OVERDUE_DAYS,
    admin_has_role,
    assign_collection_admin_if_needed_async,
    assign_collection_admins_for_overdue_loans_async,
    assign_review_admin_if_needed_async,
    is_collection_stage,
    list_admins_by_role_async,
)
from app.services.overdue_fee_config import calculate_overdue_days, calculate_penalty_by_repayment_date
from app.services.compliance import validate_cash_loan_compliance_async, normalize_cash_loan_fee_components
from app.services.scheduler import scheduler
from app.services.risk_report import (
    get_or_create_risk_report_async,
    get_user_for_risk_report_async,
    serialize_risk_report,
)
from app.services.composite_risk_report import (
    get_or_create_composite_risk_report_async,
    get_or_create_standalone_composite_risk_report_async,
    serialize_composite_risk_report,
)
from app.services.upload_storage import build_upload_url
from app.services.momo import complete_momo_transaction, create_or_get_momo_transaction_async, momo_provider
from app.services.blacklist_service import (
    add_blacklist_entry,
    blacklist_user,
    refresh_user_blacklist_status,
    remove_user_from_blacklist,
    serialize_blacklist_entry,
    upload_blacklist_entries,
)


LOAN_STATUSES = {
    "INIT",
    "REVIEWING",
    "APPROVED",
    "REJECTED",
    "WITHDRAWING",
    "DISBURSED",
    "SETTLED",
    "OVERDUE",
    "CARD_REJECTED",
}

LOAN_PAGE_PERMISSION_KEYS = (
    "applications",
    "disbursements",
    "repayments",
    "collections",
    "financials",
)

ADMIN_STATS_PERMISSION_KEYS = (
    "overview",
    "applications",
    "disbursements",
    "repayments",
    "collections",
    "financials",
)

REPAYMENT_STATS_PERMISSION_KEYS = (
    "overview",
    "repayments",
    "financials",
)

ECARD_POOL_STATUSES = {"AVAILABLE", "ASSIGNED", "EXPIRED", "VOID"}

RISK_LOCATION_KEYWORDS = tuple(
    dict.fromkeys(
        [
            "潍坊",
            "瓦房店",
            "聊城",
            "无锡",
            "宜兴",
            "日照",
            "烟台",
            "新疆",
            "长沙",
            "鹿泉区汇源街",
            "大连",
            "张家口",
            "葫芦岛",
            "内蒙古",
        ]
    )
)
RISK_LOCATION_OVERSEAS_KEYWORD = "中国大陆境外"
MAINLAND_COUNTRY_NAMES = {"中国", "中国大陆", "中华人民共和国"}
ADMIN_STATS_WS_PUSH_SECONDS = 5

_ADMIN_STATS_VERSION = 0


async def notify_admin_stats_changed():
    """标记后台统计数据已变化，并唤醒等待中的 WS 推送循环。

    :return: None
    """
    global _ADMIN_STATS_VERSION
    _ADMIN_STATS_VERSION += 1


async def wait_admin_stats_changed(last_version: int, timeout_seconds: float):
    """等待后台统计数据变更；超时则返回当前版本用于兜底推送。

    :param last_version: 上次已推送版本
    :param timeout_seconds: 超时时间（秒）
    :return: 最新版本号
    """
    if _ADMIN_STATS_VERSION != last_version:
        return _ADMIN_STATS_VERSION

    deadline = asyncio.get_running_loop().time() + max(float(timeout_seconds or 0), 0)
    while asyncio.get_running_loop().time() < deadline:
        await asyncio.sleep(0.5)
        if _ADMIN_STATS_VERSION != last_version:
            return _ADMIN_STATS_VERSION
    return _ADMIN_STATS_VERSION



def get_today_range():
    now = datetime.now()
    today_start = datetime(now.year, now.month, now.day)
    tomorrow = today_start + timedelta(days=1)
    return today_start, tomorrow

async def _get_ws_admin_by_token(db: AsyncSession, token: Optional[str]) -> Optional[Admin]:
    """通过 WebSocket Token 获取管理员。

    :param db: 异步数据库会话
    :param token: token 字符串
    :return: 管理员对象；token 非法或为空时返回 None
    """
    if not token:
        return None
    try:
        return await get_admin_by_token_async(db, token)
    except HTTPException:
        return None

def _extract_ws_token(websocket: WebSocket) -> Optional[str]:
    """提取 WebSocket 连接中的管理员 token。

    :param websocket: WebSocket 连接对象
    :return: token 字符串
    """
    query_token = websocket.query_params.get("token")
    if query_token:
        return query_token
    auth_header = websocket.headers.get("authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header.split(" ", 1)[1].strip() or None
    return None

def calculate_due_date(disbursed_at: datetime, term_days: Optional[int]):
    if not disbursed_at or not term_days:
        return None

    offset_days = max(int(term_days) - 1, 0)
    return disbursed_at + timedelta(days=offset_days)

def ensure_valid_term_days(term_days: Optional[int]) -> Optional[int]:
    if term_days is None:
        return None

    try:
        return normalize_term_days(term_days)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _set_available_credit(user: User, amount: float):
    user.available_credit_limit = round_money(max(float(amount or 0), 0))


def _extract_scalar_items(execute_result):
    """兼容真实查询结果与单元测试桩，提取标量列表。

    :param execute_result: SQLAlchemy execute 返回值或测试桩
    :return: 标量对象列表
    """
    scalars = execute_result.scalars()
    if hasattr(scalars, "all"):
        return scalars.all()
    if hasattr(scalars, "first"):
        first_item = scalars.first()
        return [] if first_item is None else [first_item]
    return []


def _ecard_amount_cents(value: float) -> int:
    """将E卡金额转换为分，避免浮点比较误差。

    :param value: 金额
    :return: 分单位金额
    """
    return int(round(round_money(value) * 100))


def _is_better_ecard_combo(candidate: list, current: Optional[list]) -> bool:
    """判断候选组合是否优于当前组合。

    :param candidate: 候选E卡列表
    :param current: 当前E卡列表
    :return: 是否优于当前组合
    """
    if current is None:
        return True
    candidate_key = (len(candidate), [item.expires_at for item in candidate], [item.id for item in candidate])
    current_key = (len(current), [item.expires_at for item in current], [item.id for item in current])
    return candidate_key < current_key


def _select_ecard_combo(candidates: list, target_face_value: float) -> list:
    """从卡池候选中选择总额精确匹配的E卡组合。

    :param candidates: 可用E卡候选列表
    :param target_face_value: 目标E卡总面额
    :return: 精确匹配的E卡组合
    """
    target_cents = _ecard_amount_cents(target_face_value)
    if target_cents <= 0:
        return []

    combos: dict[int, list] = {0: []}
    for item in candidates:
        item_cents = _ecard_amount_cents(getattr(item, "face_value", 0))
        if item_cents <= 0 or item_cents > target_cents:
            continue
        for amount, combo in list(combos.items()):
            next_amount = amount + item_cents
            if next_amount > target_cents:
                continue
            next_combo = sorted(combo + [item], key=lambda card: (card.expires_at, card.id))
            if _is_better_ecard_combo(next_combo, combos.get(next_amount)):
                combos[next_amount] = next_combo

    return combos.get(target_cents, [])


async def _find_ready_fee_extension_orders(db: AsyncSession, source_loan: Loan):
    result = await db.execute(
        select(Loan)
        .where(
            Loan.user_id == source_loan.user_id,
            Loan.is_extension_fee_order.is_(True),
            Loan.extension_source_loan_id == source_loan.id,
            Loan.extension_used_at.is_(None),
            Loan.status == "DISBURSED",
        )
        .order_by(Loan.disbursed_at.asc(), Loan.id.asc())
    )
    return result.scalars().all()


def _resolve_order_submit_time(loan: Loan) -> Optional[datetime]:
    """解析订单真实下单时间。

    :param loan: 订单对象
    :return: 用户提交信用下单的时间，取不到时返回 None
    """
    events = getattr(getattr(loan, "owner", None), "events", None) or []
    matched_events = [
        event
        for event in events
        if getattr(event, "loan_id", None) == getattr(loan, "id", None)
        and getattr(event, "event_type", None) == "ORDER_SUBMIT"
        and getattr(event, "created_at", None)
    ]
    if not matched_events:
        return None
    # 同一订单理论上只有一次下单事件；若历史数据有重复，按最近一次真实提交为准。
    return max(event.created_at for event in matched_events)


def serialize_admin_user(admin: Admin, current_admin: Optional[Admin] = None):
    roles = parse_admin_roles(getattr(admin, "roles", None))
    return {
        "id": admin.id,
        "username": admin.username,
        "roles": roles,
        "permissions": resolve_admin_permissions(admin),
        "created_at": admin.created_at,
        "updated_at": admin.updated_at,
        "is_active": bool(getattr(admin, "is_active", True)),
        "is_current": bool(current_admin and admin.id == current_admin.id),
    }

def resolve_roles_and_permissions(roles_input, permissions_input):
    roles = parse_admin_roles(roles_input)
    if roles:
        return roles, resolve_permissions_from_roles(roles)

    permissions = [
        item for item in parse_admin_permissions(permissions_input)
        if item in ALL_ADMIN_PERMISSION_KEYS
    ]
    if permissions:
        # 兼容旧请求：仅传页面权限时，按页面权限落库
        return [], permissions
    return [], []

def ensure_admin_page_permission(current_admin: Admin, permission_key: str):
    if not admin_has_permission(current_admin, permission_key):
        raise HTTPException(status_code=403, detail="无权访问当前页面")

def ensure_admin_permission(current_admin: Admin, permission_key: str, detail: str = "无权执行当前操作"):
    """校验后台用户是否拥有指定操作权限。

    :param current_admin: 当前登录管理员
    :param permission_key: 权限标识
    :param detail: 无权限时的提示文案
    :return: None
    """
    if not admin_has_permission(current_admin, permission_key):
        raise HTTPException(status_code=403, detail=detail)

def ensure_any_admin_page_permission(current_admin: Admin, permission_keys):
    if any(admin_has_permission(current_admin, item) for item in permission_keys):
        return
    raise HTTPException(status_code=403, detail="无权访问当前页面")

def resolve_loan_scope_permission(
    scope: Optional[str],
    due_date_preset: Optional[str],
    status: Optional[str] = None,
):
    if scope == "REVIEWING":
        return "applications"
    if scope == "WITHDRAWING":
        return "disbursements"
    if scope == "FINANCE":
        return "financials"
    if scope == "OVERDUE":
        return "collections"
    if scope == "DUE_TODAY" or due_date_preset in {"TODAY", "TOMORROW"}:
        return "repayments"
    if scope == "REPAYMENTS":
        return "repayments"

    if status in {"REVIEWING", "APPROVED", "REJECTED"}:
        return "applications"
    if status == "WITHDRAWING":
        return "disbursements"
    if status in {"DISBURSED", "SETTLED"}:
        return "repayments"
    if status == "OVERDUE":
        return "collections"

    return None

def current_admin_roles(current_admin: Admin):
    return parse_admin_roles(getattr(current_admin, "roles", None))

def is_super_admin(current_admin: Admin) -> bool:
    return "ADMIN" in current_admin_roles(current_admin)

def ensure_stage_access_for_admin(current_admin: Admin, loan: Loan):
    roles = current_admin_roles(current_admin)
    if "ADMIN" in roles:
        return
    if is_collection_stage(loan):
        if "COLLECTION" in roles and int(loan.collection_admin_id or 0) == int(current_admin.id):
            return
        raise HTTPException(status_code=403, detail="当前订单已转催收，仅限负责催收员处理")

    if "REVIEW" in roles and int(loan.review_admin_id or 0) == int(current_admin.id):
        return
    if "FINANCE" in roles and loan.status in {"DISBURSED", "OVERDUE", "SETTLED", "WITHDRAWING"}:
        return

    raise HTTPException(status_code=403, detail="当前订单未分配给你处理")


def _append_location_text(parts: list[str], *values: Optional[str]) -> None:
    """追加非空地址文本。

    :param parts: 地址片段列表
    :param values: 待追加的地址字段
    :return: 无返回值
    """
    for value in values:
        text = (value or "").strip()
        if text:
            parts.append(text)


def _safe_loaded_events(user: Optional[User]) -> list[UserEvent]:
    """读取已预加载的用户访问日志。

    :param user: 用户对象
    :return: 已加载的访问日志列表
    """
    if not user:
        return []
    # User.events 使用 noload，直接从 __dict__ 读取可避免列表接口触发额外查询。
    events = getattr(user, "__dict__", {}).get("events")
    return events or []


def _is_overseas_country(country: Optional[str]) -> bool:
    """判断国家字段是否属于中国大陆境外。

    :param country: 国家或地区名称
    :return: 是否命中境外风险
    """
    text = (country or "").strip()
    if not text:
        return False
    if text in MAINLAND_COUNTRY_NAMES:
        return False
    return "中国" not in text or "香港" in text or "澳门" in text or "台湾" in text


def resolve_user_location_risk(user: Optional[User]) -> dict:
    """根据用户GPS与IP解析地址判断是否命中风险地区。

    :param user: 用户对象
    :return: 风险命中结果
    """
    parts: list[str] = []
    overseas_hit = False

    if user:
        _append_location_text(
            parts,
            getattr(user, "location_province", None),
            getattr(user, "location_city", None),
            getattr(user, "location_district", None),
            getattr(user, "location_street", None),
            getattr(user, "location_address", None),
            getattr(user, "id_address", None),
        )

    for event in _safe_loaded_events(user):
        _append_location_text(
            parts,
            getattr(event, "ip_province", None),
            getattr(event, "ip_city", None),
            getattr(event, "ip_district", None),
            getattr(event, "ip_detail", None),
            getattr(event, "lon_lat_province", None),
            getattr(event, "lon_lat_city", None),
            getattr(event, "lon_lat_district", None),
            getattr(event, "lon_lat_detail", None),
        )
        overseas_hit = overseas_hit or _is_overseas_country(getattr(event, "ip_country", None))
        overseas_hit = overseas_hit or _is_overseas_country(getattr(event, "lon_lat_country", None))

    combined_text = " ".join(parts)
    keywords = [keyword for keyword in RISK_LOCATION_KEYWORDS if keyword in combined_text]
    if overseas_hit or RISK_LOCATION_OVERSEAS_KEYWORD in combined_text:
        keywords.append(RISK_LOCATION_OVERSEAS_KEYWORD)

    unique_keywords = list(dict.fromkeys(keywords))
    return {
        "hit": bool(unique_keywords),
        "keywords": unique_keywords,
        "detail": "命中风险位置：" + "、".join(unique_keywords) if unique_keywords else "",
    }


def serialize_loan(loan: Loan):
    owner_loans = getattr(getattr(loan, "owner", None), "loans", None) or []
    loan.relend_count = get_relend_count(owner_loans, current_loan_id=loan.id)
    loan.relend_label = get_relend_label(owner_loans, current_loan_id=loan.id)
    loan.latest_settled_loan = get_latest_normal_settled_loan(owner_loans, current_loan_id=loan.id)
    payload = serialize_loan_snapshot(loan, include_user=True)
    payload["ordered_at"] = _resolve_order_submit_time(loan)
    location_risk = resolve_user_location_risk(getattr(loan, "owner", None))
    payload["user_location_risk_hit"] = location_risk["hit"]
    payload["user_location_risk_keywords"] = location_risk["keywords"]
    payload["user_location_risk_detail"] = location_risk["detail"]
    payload["review_admin_id"] = loan.review_admin_id
    payload["review_admin_name"] = None
    payload["collection_admin_id"] = loan.collection_admin_id
    payload["collection_admin_name"] = None
    payload["collection_transferred_at"] = loan.collection_transferred_at

    if loan.review_admin_id:
        reviewer = getattr(loan, "review_admin", None)
        payload["review_admin_name"] = reviewer.username if reviewer else None
    if loan.collection_admin_id:
        collector = getattr(loan, "collection_admin", None)
        payload["collection_admin_name"] = collector.username if collector else None
    return payload

def serialize_user_summary(user: User):
    latest_loan = max(user.loans, key=lambda item: item.id) if user.loans else None
    if latest_loan:
        latest_loan.relend_count = get_relend_count(user.loans, current_loan_id=latest_loan.id)
        latest_loan.relend_label = get_relend_label(user.loans, current_loan_id=latest_loan.id)
        latest_loan.latest_settled_loan = get_latest_normal_settled_loan(user.loans, current_loan_id=latest_loan.id)
    loan_snapshot = serialize_loan_snapshot(latest_loan) if latest_loan else None
    disbursed_loans = [item for item in (user.loans or []) if getattr(item, "disbursed_at", None)]
    first_deal_loan = min(disbursed_loans, key=lambda item: item.id) if disbursed_loans else None
    latest_deal_loan = max(disbursed_loans, key=lambda item: item.id) if disbursed_loans else None
    source_channel = user.source_channel
    return {
        "id": user.id,
        "phone": mask_secret(user.phone, left=0, right=4),
        "name": user.name,
        "id_card_num": mask_secret(user.id_card_num, left=6, right=1),
        "id_card_front_image_url": build_upload_url(getattr(user, "id_card_front_image", None)),
        "id_card_back_image_url": build_upload_url(getattr(user, "id_card_back_image", None)),
        "face_image_url": build_upload_url(getattr(user, "face_image", None)),
        "face_auth_status": user.face_auth_status,
        "approved_limit": user.approved_limit,
        "available_credit_limit": round_money(getattr(user, "available_credit_limit", 0)),
        "overdue_credit_locked": bool(getattr(user, "overdue_credit_locked", False)),
        "blacklist_hit": bool(getattr(user, "blacklist_hit", False)),
        "blacklist_reason": getattr(user, "blacklist_reason", None),
        "blacklist_checked_at": getattr(user, "blacklist_checked_at", None),
        "location_risk_blocked": bool(getattr(user, "location_risk_blocked", False)),
        "location_risk_reason": getattr(user, "location_risk_reason", None),
        "location_risk_at": getattr(user, "location_risk_at", None),
        "risk_list_hit": bool(getattr(user, "risk_list_hit", False)),
        "risk_list_source": getattr(user, "risk_list_source", None),
        "risk_list_reason": getattr(user, "risk_list_reason", None),
        "risk_list_checked_at": getattr(user, "risk_list_checked_at", None),
        "created_at": user.created_at,
        "last_login_at": user.last_login_at,
        "application_submitted_at": user.application_submitted_at,
        "current_loan_id": latest_loan.id if latest_loan else None,
        "current_loan_status": latest_loan.status if latest_loan else None,
        "current_credit_limit": loan_snapshot["credit_limit"] if loan_snapshot else None,
        "current_term_days": loan_snapshot["term_days"] if loan_snapshot else None,
        "current_due_date": loan_snapshot["due_date"] if loan_snapshot else None,
        "current_fee_rate": loan_snapshot["fee_rate"] if loan_snapshot else None,
        "current_fee_amount": loan_snapshot["fee_amount"] if loan_snapshot else None,
        "current_interest_amount": loan_snapshot["interest_amount"] if loan_snapshot else None,
        "current_guarantee_fee_amount": loan_snapshot["guarantee_fee_amount"] if loan_snapshot else None,
        "current_penalty_amount": loan_snapshot["penalty_amount"] if loan_snapshot else None,
        "current_blacklist_hit": bool(getattr(user, "blacklist_hit", False)),
        "current_risk_list_hit": bool(getattr(user, "risk_list_hit", False)),
        "first_disbursed_at": getattr(first_deal_loan, "disbursed_at", None),
        "first_deal_amount": round_money(getattr(first_deal_loan, "product_total_price", 0)) if first_deal_loan else None,
        "latest_disbursed_at": getattr(latest_deal_loan, "disbursed_at", None),
        "latest_deal_amount": round_money(getattr(latest_deal_loan, "product_total_price", 0)) if latest_deal_loan else None,
        "source_channel_name": source_channel.channel_name if source_channel else None,
        "source_channel_sales_name": source_channel.sales_name if source_channel else None,
        "channel_bound_at": user.channel_bound_at,
        "last_channel_visit_at": user.last_channel_visit_at,
    }

def serialize_user_detail(user: User, events: Optional[list[UserEvent]] = None):
    latest_loan = max(user.loans, key=lambda item: item.id) if user.loans else None
    if latest_loan:
        latest_loan.relend_count = get_relend_count(user.loans, current_loan_id=latest_loan.id)
        latest_loan.relend_label = get_relend_label(user.loans, current_loan_id=latest_loan.id)
        latest_loan.latest_settled_loan = get_latest_normal_settled_loan(user.loans, current_loan_id=latest_loan.id)
    event_list = sorted(events or [], key=lambda item: item.created_at, reverse=True)
    source_channel = user.source_channel
    # 首次成交口径：优先第一条非拒绝订单，取不到时回退第一条订单。
    sorted_loans = sorted(user.loans or [], key=lambda item: (item.created_at or datetime.min, item.id or 0))
    first_deal_loan = next((item for item in sorted_loans if item.status != "REJECTED"), None)
    if first_deal_loan is None and sorted_loans:
        first_deal_loan = sorted_loans[0]
    if first_deal_loan:
        first_deal_loan.relend_count = get_relend_count(user.loans, current_loan_id=first_deal_loan.id)
        first_deal_loan.relend_label = get_relend_label(user.loans, current_loan_id=first_deal_loan.id)
        first_deal_loan.latest_settled_loan = get_latest_normal_settled_loan(user.loans, current_loan_id=first_deal_loan.id)
    return {
        "id": user.id,
        "phone": user.phone,
        "name": user.name,
        "id_card_num": user.id_card_num,
        "id_address": user.id_address,
        "id_expiry": user.id_expiry,
        "id_card_front_image_url": build_upload_url(getattr(user, "id_card_front_image", None)),
        "id_card_back_image_url": build_upload_url(getattr(user, "id_card_back_image", None)),
        "face_image_url": build_upload_url(getattr(user, "face_image", None)),
        "approved_limit": user.approved_limit,
        "available_credit_limit": round_money(getattr(user, "available_credit_limit", 0)),
        "overdue_credit_locked": bool(getattr(user, "overdue_credit_locked", False)),
        "blacklist_hit": bool(getattr(user, "blacklist_hit", False)),
        "blacklist_reason": getattr(user, "blacklist_reason", None),
        "blacklist_checked_at": getattr(user, "blacklist_checked_at", None),
        "risk_list_hit": bool(getattr(user, "risk_list_hit", False)),
        "risk_list_source": getattr(user, "risk_list_source", None),
        "risk_list_reason": getattr(user, "risk_list_reason", None),
        "risk_list_checked_at": getattr(user, "risk_list_checked_at", None),
        "emergency_contact1_name": user.emergency_contact1_name,
        "emergency_contact1_relation": user.emergency_contact1_relation,
        "emergency_contact1_phone": user.emergency_contact1_phone,
        "emergency_contact2_name": user.emergency_contact2_name,
        "emergency_contact2_relation": user.emergency_contact2_relation,
        "emergency_contact2_phone": user.emergency_contact2_phone,
        "location_latitude": user.location_latitude,
        "location_longitude": user.location_longitude,
        "location_accuracy": user.location_accuracy,
        "location_address": user.location_address,
        "location_province": user.location_province,
        "location_city": user.location_city,
        "location_district": user.location_district,
        "location_street": user.location_street,
        "location_source": user.location_source,
        "location_updated_at": user.location_updated_at,
        "location_risk_blocked": bool(getattr(user, "location_risk_blocked", False)),
        "location_risk_reason": getattr(user, "location_risk_reason", None),
        "location_risk_at": getattr(user, "location_risk_at", None),
        "face_auth_status": user.face_auth_status,
        "face_auth_at": user.face_auth_at,
        "last_login_at": user.last_login_at,
        "ocr_submitted_at": user.ocr_submitted_at,
        "application_submitted_at": user.application_submitted_at,
        "source_channel_name": source_channel.channel_name if source_channel else None,
        "source_channel_sales_name": source_channel.sales_name if source_channel else None,
        "channel_bound_at": user.channel_bound_at,
        "last_channel_visit_at": user.last_channel_visit_at,
        "created_at": user.created_at,
        "latest_loan": serialize_loan_snapshot(latest_loan, include_ledger=True) if latest_loan else None,
        "first_deal_loan": serialize_loan_snapshot(first_deal_loan, include_ledger=True) if first_deal_loan else None,
        "events": [
            {
                "id": event.id,
                "loan_id": event.loan_id,
                "actor_type": event.actor_type,
                "operator_name": event.operator_name,
                "event_type": event.event_type,
                "title": event.title,
                "detail": event.detail,
                "ip": event.ip or "",
                "ip_country": event.ip_country or "",
                "ip_province": event.ip_province or "",
                "ip_city": event.ip_city or "",
                "ip_district": event.ip_district or "",
                "ip_detail": event.ip_detail or "",
                "lon_lat": event.lon_lat or "",
                "lon_lat_province": event.lon_lat_province or "",
                "lon_lat_city": event.lon_lat_city or "",
                "lon_lat_district": event.lon_lat_district or "",
                "lon_lat_detail": event.lon_lat_detail or "",
                "created_at": event.created_at,
            }
            for event in event_list
        ],
    }


async def _get_user_ip_audit(db: AsyncSession, current_admin: Admin, user_id: int):
    ensure_any_admin_page_permission(current_admin, ("users", "applications", "disbursements", "repayments", "collections"))
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    events = (
        await db.execute(
            select(UserEvent)
            .where(UserEvent.user_id == user_id, UserEvent.ip != "")
            .order_by(UserEvent.created_at.desc(), UserEvent.id.desc())
        )
    ).scalars().all()
    seen = set()
    items = []
    for event in events:
        ip = (event.ip or "").strip()
        if not ip or ip in seen:
            continue
        seen.add(ip)
        items.append(
            {
                "ip": ip,
                "operation": event.title or event.event_type,
                "detail": event.detail,
                "created_at": event.created_at,
                "country": event.ip_country,
                "province": event.ip_province,
                "city": event.ip_city,
                "district": event.ip_district,
                "address": event.ip_detail,
            }
        )
    return {"user_id": user.id, "name": user.name, "phone": mask_secret(user.phone, left=0, right=4), "items": items}

def serialize_channel(channel: Channel, advisor: Optional[Admin] = None):
    payload = build_channel_metrics(channel)
    payload["invite_code"] = channel.invite_code
    payload["daily_invite_code"] = build_daily_channel_invite_code(channel)
    payload["admin_user_id"] = channel.admin_user_id
    payload["admin_user_name"] = advisor.username if advisor else None
    payload["disbursement_mode"] = normalize_channel_disbursement_mode(getattr(channel, "disbursement_mode", None))
    payload["review_mode"] = normalize_channel_review_mode(getattr(channel, "review_mode", None))
    return payload

def _generate_channel_invite_code(length: int = 16) -> str:
    """生成渠道邀请码。

    :param length: 邀请码长度
    :return: 邀请码
    """
    alphabet = string.ascii_lowercase + string.digits
    while True:
        invite_code = "".join(secrets.choice(alphabet) for _ in range(length))
        if any(char.isalpha() for char in invite_code) and any(char.isdigit() for char in invite_code):
            return invite_code


async def _generate_unique_channel_invite_code(db: AsyncSession, length: int = 16) -> str:
    """生成数据库内唯一的渠道邀请码。

    :param db: 异步数据库会话
    :param length: 邀请码长度
    :return: 唯一邀请码
    """
    for _ in range(10):
        invite_code = _generate_channel_invite_code(length)
        exists = (await db.execute(select(Channel).where(Channel.invite_code == invite_code))).scalar_one_or_none()
        if not exists:
            return invite_code
    raise HTTPException(status_code=500, detail="生成渠道邀请码失败，请稍后重试")


def _is_business_consultant(admin: Optional[Admin]) -> bool:
    """判断后台用户是否为业务顾问角色（严格匹配，不包含 ADMIN 兜底）。

    :param admin: 后台用户对象
    :return: 是否包含 BUSINESS_CONSULTANT 角色
    """
    if admin is None:
        return False
    return "BUSINESS_CONSULTANT" in parse_admin_roles(getattr(admin, "roles", None))

def _is_only_business_consultant(admin: Optional[Admin]) -> bool:
    """判断后台用户是否为仅业务顾问角色。

    :param admin: 后台用户对象
    :return: 是否有且仅有 BUSINESS_CONSULTANT 角色
    """
    if admin is None:
        return False
    roles = parse_admin_roles(getattr(admin, "roles", None))
    return len(roles) == 1 and roles[0] == "BUSINESS_CONSULTANT"

def apply_business_consultant_user_summary_status(user_summary: dict, admin: Optional[Admin]) -> dict:
    """按业务顾问规则调整用户列表状态展示值。

    :param user_summary: 用户列表摘要数据
    :param admin: 当前登录后台用户
    :return: 调整后的用户列表摘要数据
    """
    if _is_only_business_consultant(admin) and user_summary.get("first_disbursed_at"):
        payload = dict(user_summary)
        payload["current_loan_status"] = "FIRST_BORROW"
        return payload
    return user_summary


async def _resolve_business_advisor_by_id(db: AsyncSession, admin_user_id: int):
    advisor = (await db.execute(select(Admin).where(Admin.id == admin_user_id))).scalar_one_or_none()
    if not _is_business_consultant(advisor):
        raise HTTPException(status_code=400, detail="请选择角色为业务顾问的后台用户")
    return advisor

def round_money(value: Optional[float]) -> float:
    return round(float(value or 0), 2)

def mask_secret(value: Optional[str], left: int = 3, right: int = 2) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) <= left + right:
        return "*" * len(text)
    return f"{text[:left]}{'*' * (len(text) - left - right)}{text[-right:]}"

def resolve_product_payment_amount(ecard_face_value: float, rights_price: float, payment_amount: Optional[float] = None):
    if payment_amount is not None and float(payment_amount) > 0:
        return round_money(payment_amount)
    return round_money(float(ecard_face_value or 0) + float(rights_price or 0))

def serialize_product(product: Product):
    rights_detail = None
    if getattr(product, "rights_detail_json", None):
        try:
            rights_detail = json.loads(product.rights_detail_json)
        except (TypeError, ValueError):
            rights_detail = None
    fee_components = None
    if getattr(product, "fee_components_json", None):
        try:
            fee_components = json.loads(product.fee_components_json)
        except (TypeError, ValueError):
            fee_components = None
    ratios = normalize_installment_ratios(
        getattr(product, "installment_ratios_json", None),
        getattr(product, "installment_count", 1),
    )
    return {
        "id": product.id,
        "name": product.name,
        "ecard_face_value": round_money(product.ecard_face_value),
        "rights_price": round_money(product.rights_price),
        "rights_title": product.rights_title,
        "rights_desc": product.rights_desc,
        "rights_detail": rights_detail,
        "term_days": product.term_days,
        "payment_amount": round_money(product.payment_amount),
        "nominal_loan_amount": round_money(getattr(product, "nominal_loan_amount", 0) or product.payment_amount),
        "upfront_fee_rate": float(getattr(product, "upfront_fee_rate", 0.4) or 0.4),
        "upfront_fee_amount": round_money(
            (getattr(product, "nominal_loan_amount", 0) or product.payment_amount or 0)
            * (getattr(product, "upfront_fee_rate", 0.4) or 0.4)
        ) if getattr(product, "product_type", None) == "CASH_LOAN" else 0,
        "actual_disbursement_amount": round_money(
            max(
                (getattr(product, "nominal_loan_amount", 0) or product.payment_amount or 0)
                - (getattr(product, "nominal_loan_amount", 0) or product.payment_amount or 0)
                * (getattr(product, "upfront_fee_rate", 0.4) or 0.4),
                0,
            )
        ) if getattr(product, "product_type", None) == "CASH_LOAN" else 0,
        "fee_components": fee_components,
        "interest_start_day": int(getattr(product, "interest_start_day", 1) or 1),
        "repayment_due_day": int(getattr(product, "repayment_due_day", 7) or 7),
        "installment_count": int(getattr(product, "installment_count", 1) or 1),
        "installment_ratios": ratios,
        "daily_overdue_fee": round_money(getattr(product, "daily_overdue_fee", 10) or 10),
        "borrower_type": getattr(product, "borrower_type", None) or "ALL",
        "product_type": getattr(product, "product_type", None) or "ECARD_RIGHTS",
        "is_active": bool(product.is_active),
        "created_at": product.created_at,
        "updated_at": product.updated_at,
    }

def serialize_ecard_pool_item(item: EcardPool):
    recipient_phone = getattr(item, "recipient_phone", None)
    secret_copied_at = getattr(item, "secret_copied_at", None)
    return {
        "id": item.id,
        "account": mask_secret(item.account, left=4, right=4),
        "password": mask_secret(item.password, left=1, right=1),
        "face_value": round_money(item.face_value),
        "expires_at": item.expires_at,
        "status": item.status,
        "loan_id": item.loan_id,
        "recipient_phone": recipient_phone,
        "secret_copied_at": secret_copied_at,
        "note": item.note,
        "assigned_at": item.assigned_at,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }

def _extract_ecard_pool_id_from_copy_detail(detail: Optional[str]) -> Optional[int]:
    """从用户复制卡密事件详情中提取卡池ID。

    :param detail: 用户事件详情
    :return: 卡池ID，无法解析时返回 None
    """
    match = re.search(r"(?:^|；)ecard_pool_id=(\d+)(?:；|$)", detail or "")
    return int(match.group(1)) if match else None

async def _build_ecard_pool_stats(db: AsyncSession) -> dict:
    """构建卡池管理页顶部统计卡片数据。

    :param db: 异步数据库会话
    :return: 卡池统计字典
    """
    today_start, tomorrow = get_today_range()
    total_count, total_amount = (
        await db.execute(select(func.count(EcardPool.id), func.coalesce(func.sum(EcardPool.face_value), 0)))
    ).one()
    cumulative_assigned_count, cumulative_assigned_amount = (
        await db.execute(
            select(func.count(EcardPool.id), func.coalesce(func.sum(EcardPool.face_value), 0)).where(
                EcardPool.assigned_at.isnot(None),
            )
        )
    ).one()
    available_count, available_amount = (
        await db.execute(
            select(func.count(EcardPool.id), func.coalesce(func.sum(EcardPool.face_value), 0)).where(
                EcardPool.status == "AVAILABLE",
                EcardPool.expires_at >= tomorrow,
            )
        )
    ).one()
    today_stock_in_count, today_stock_in_amount = (
        await db.execute(
            select(func.count(EcardPool.id), func.coalesce(func.sum(EcardPool.face_value), 0)).where(
                EcardPool.created_at >= today_start,
                EcardPool.created_at < tomorrow,
            )
        )
    ).one()
    today_assigned_count, today_assigned_amount = (
        await db.execute(
            select(func.count(EcardPool.id), func.coalesce(func.sum(EcardPool.face_value), 0)).where(
                EcardPool.assigned_at >= today_start,
                EcardPool.assigned_at < tomorrow,
            )
        )
    ).one()
    return {
        "pool_total_count": int(total_count or 0),
        "pool_total_amount": round_money(total_amount),
        "cumulative_assigned_count": int(cumulative_assigned_count or 0),
        "cumulative_assigned_amount": round_money(cumulative_assigned_amount),
        "available_count": int(available_count or 0),
        "available_amount": round_money(available_amount),
        "today_stock_in_count": int(today_stock_in_count or 0),
        "today_stock_in_amount": round_money(today_stock_in_amount),
        "today_assigned_count": int(today_assigned_count or 0),
        "today_assigned_amount": round_money(today_assigned_amount),
    }

async def _attach_ecard_pool_display_fields(db: AsyncSession, items: list[EcardPool]) -> None:
    """给卡池列表项补充领取人手机号和首次复制密码时间。

    :param db: 异步数据库会话
    :param items: 当前页卡池列表
    :return: None
    """
    if not items:
        return

    loan_ids = [item.loan_id for item in items if item.loan_id]
    if loan_ids:
        owner_rows = (
            await db.execute(
                select(Loan.id, User.phone)
                .join(User, User.id == Loan.user_id)
                .where(Loan.id.in_(loan_ids))
            )
        ).all()
        phone_by_loan_id = {loan_id: phone for loan_id, phone in owner_rows}
    else:
        phone_by_loan_id = {}

    pool_ids = [item.id for item in items]
    copy_filters = [
        UserEvent.event_type == "USER_ECARD_SECRET_COPIED",
        UserEvent.detail.like("%field=password%"),
        UserEvent.detail.like("%ecard_pool_id=%"),
    ]
    if loan_ids:
        # 复制事件与订单绑定，先按当前页订单收窄范围，避免卡池列表扫描过多用户事件。
        copy_filters.append(UserEvent.loan_id.in_(loan_ids))
    copy_rows = (
        await db.execute(
            select(UserEvent.detail, UserEvent.created_at)
            .where(*copy_filters)
            .order_by(UserEvent.created_at.asc())
        )
    ).all()
    pool_id_set = set(pool_ids)
    copied_at_by_pool_id = {}
    for detail, created_at in copy_rows:
        pool_id = _extract_ecard_pool_id_from_copy_detail(detail)
        if pool_id in pool_id_set and pool_id not in copied_at_by_pool_id:
            copied_at_by_pool_id[pool_id] = created_at

    for item in items:
        item.recipient_phone = phone_by_loan_id.get(item.loan_id)
        item.secret_copied_at = copied_at_by_pool_id.get(item.id)

def apply_loan_scope(query, scope: Optional[str]):
    overdue_days_expr = func.datediff(func.current_date(), func.date(Loan.due_date))
    today_start, tomorrow = get_today_range()

    if scope == "REVIEWING":
        return query.filter(Loan.status.in_(["REVIEWING", "APPROVED", "REJECTED"]))
    if scope == "WITHDRAWING":
        return query.filter(Loan.status == "WITHDRAWING")
    if scope == "FINANCE":
        return query.filter(Loan.status.in_(["DISBURSED", "OVERDUE"]))
    if scope == "DUE_TODAY":
        return query.filter(
            Loan.status.in_(["DISBURSED", "OVERDUE"]),
            Loan.due_date >= today_start,
            Loan.due_date < tomorrow,
        )
    if scope == "OVERDUE":
        return query.filter(
            Loan.status == "OVERDUE",
            Loan.due_date.isnot(None),
            overdue_days_expr > COLLECTION_TRANSFER_OVERDUE_DAYS,
        )
    if scope == "REPAYMENTS":
        return query.filter(
            or_(
                Loan.status.in_(["DISBURSED", "SETTLED"]),
                (
                    (Loan.status == "OVERDUE")
                    & Loan.due_date.isnot(None)
                    & (overdue_days_expr <= COLLECTION_TRANSFER_OVERDUE_DAYS)
                ),
            )
        )
    return query

def build_loan_scope_filters(scope: Optional[str]):
    overdue_days_expr = func.datediff(func.current_date(), func.date(Loan.due_date))
    today_start, tomorrow = get_today_range()

    if scope == "REVIEWING":
        return [Loan.status.in_(["REVIEWING", "APPROVED", "REJECTED"])]
    if scope == "WITHDRAWING":
        return [Loan.status == "WITHDRAWING"]
    if scope == "FINANCE":
        return [Loan.status.in_(["DISBURSED", "OVERDUE"])]
    if scope == "DUE_TODAY":
        return [
            Loan.status.in_(["DISBURSED", "OVERDUE"]),
            Loan.due_date >= today_start,
            Loan.due_date < tomorrow,
        ]
    if scope == "OVERDUE":
        return [
            Loan.status == "OVERDUE",
            Loan.due_date.isnot(None),
            overdue_days_expr > COLLECTION_TRANSFER_OVERDUE_DAYS,
        ]
    if scope == "REPAYMENTS":
        return [
            or_(
                Loan.status.in_(["DISBURSED", "SETTLED"]),
                (
                    (Loan.status == "OVERDUE")
                    & Loan.due_date.isnot(None)
                    & (overdue_days_expr <= COLLECTION_TRANSFER_OVERDUE_DAYS)
                ),
            )
        ]
    return []

def get_overdue_days_expr():
    return func.greatest(func.datediff(func.current_date(), func.date(Loan.due_date)), 1)


def _normalize_actual_repayment_date(value: Optional[date]) -> date:
    """规范化实际还款日期，未填写时按当天处理。

    :param value: 实际还款日期
    :return: 规范化后的日期
    """
    return value or date.today()


def split_extra_fee_for_penalty(extra_fee_amount: float, unpaid_penalty_amount: float) -> dict:
    """将额外收款拆分为逾期费冲抵和其他费用。

    :param extra_fee_amount: 本次额外收款
    :param unpaid_penalty_amount: 当前未结清逾期费
    :return: 拆分结果
    """
    penalty_paid_now = round(min(float(extra_fee_amount or 0), float(unpaid_penalty_amount or 0)), 2)
    other_fee_amount = round(max(float(extra_fee_amount or 0) - penalty_paid_now, 0), 2)
    return {
        "penalty_paid_now": penalty_paid_now,
        "other_fee_amount": other_fee_amount,
    }

def get_loan_operating_metrics(loan: Loan):
    ledger = get_loan_ledger_snapshot(loan)
    return ledger["summary"]

def round_cash_amount(value: Optional[float]) -> float:
    return round(float(value or 0), 2)


def build_recent_insight_charts(issued_loans, today_start: datetime, days: int = 7):
    """生成洞察看板近 N 天成交和发卡趋势。

    :param issued_loans: 已发卡订单集合
    :param today_start: 今日零点
    :param days: 统计天数
    :return: 趋势图数据列表
    """
    day_starts = [today_start - timedelta(days=offset) for offset in range(days - 1, -1, -1)]
    order_counts = {item.date(): 0 for item in day_starts}
    ecard_amounts = {item.date(): 0.0 for item in day_starts}

    for loan in issued_loans:
        disbursed_at = getattr(loan, "disbursed_at", None)
        if not disbursed_at:
            continue
        disbursed_date = disbursed_at.date()
        if disbursed_date not in order_counts:
            continue
        order_counts[disbursed_date] += 1
        ecard_amounts[disbursed_date] += float(getattr(loan, "ecard_face_value", 0) or getattr(loan, "credit_limit", 0) or 0)

    points = [
        {
            "date": day.date().isoformat(),
            "label": f"{day.month}/{day.day}",
            "order_count": int(order_counts[day.date()]),
            "ecard_amount": round_cash_amount(ecard_amounts[day.date()]),
        }
        for day in day_starts
    ]
    return [
        {
            "key": "daily_order_count",
            "title": "近7天每天成交单数",
            "value_type": "count",
            "points": [
                {"date": item["date"], "label": item["label"], "value": item["order_count"]}
                for item in points
            ],
        },
        {
            "key": "daily_ecard_amount",
            "title": "近7天每天E卡发放金额",
            "value_type": "currency",
            "points": [
                {"date": item["date"], "label": item["label"], "value": item["ecard_amount"]}
                for item in points
            ],
        },
    ]

async def build_project_cash_insights(db: AsyncSession, loans, today_start: datetime, tomorrow: datetime):
    ordered_statuses = {"WITHDRAWING", "DISBURSED", "OVERDUE", "SETTLED"}
    issued_statuses = {"DISBURSED", "OVERDUE", "SETTLED"}

    order_loans = [loan for loan in loans if loan.status in ordered_statuses and float(loan.credit_limit or loan.ecard_face_value or 0) > 0]
    issued_loans = [loan for loan in order_loans if loan.status in issued_statuses]
    overdue_loans = [loan for loan in issued_loans if loan.status == "OVERDUE"]
    normal_outstanding_loans = [loan for loan in issued_loans if loan.status == "DISBURSED"]

    total_users = (await db.scalar(select(func.count(User.id)))) or 0
    today_new_users = (
        await db.scalar(
            select(func.count(User.id)).where(User.created_at >= today_start, User.created_at < tomorrow)
        )
    ) or 0

    today_order_count = (
        await db.scalar(
            select(func.count(func.distinct(UserEvent.loan_id))).where(
                UserEvent.event_type == "ORDER_SUBMIT",
                UserEvent.loan_id.isnot(None),
                UserEvent.created_at >= today_start,
                UserEvent.created_at < tomorrow,
            )
        )
    ) or 0

    today_received_amount = (
        await db.scalar(
            select(func.coalesce(func.sum(LoanTransaction.amount), 0)).where(
                LoanTransaction.transaction_type.in_(["REPAYMENT", "SETTLEMENT"]),
                LoanTransaction.created_at >= today_start,
                LoanTransaction.created_at < tomorrow,
            )
        )
    ) or 0
    today_other_fee_amount = (
        await db.scalar(
            select(func.coalesce(func.sum(LoanTransaction.amount), 0)).where(
                LoanTransaction.transaction_type == "OTHER_FEE",
                LoanTransaction.created_at >= today_start,
                LoanTransaction.created_at < tomorrow,
            )
        )
    ) or 0

    today_issued_loans = [
        loan
        for loan in issued_loans
        if loan.disbursed_at and today_start <= loan.disbursed_at < tomorrow
    ]

    total_ecard_issued = round_cash_amount(sum(float(loan.ecard_face_value or loan.credit_limit or 0) for loan in issued_loans))
    today_ecard_issued = round_cash_amount(
        sum(float(loan.ecard_face_value or loan.credit_limit or 0) for loan in today_issued_loans)
    )

    total_rights_cost = round_cash_amount(sum(float(loan.rights_price or 0) * 0.04 for loan in issued_loans))
    today_rights_cost = round_cash_amount(sum(float(loan.rights_price or 0) * 0.04 for loan in today_issued_loans))

    order_total_amount = round_cash_amount(sum(calculate_total_repayment_amount(loan) for loan in issued_loans))
    today_order_total_amount = round_cash_amount(sum(calculate_total_repayment_amount(loan) for loan in today_issued_loans))

    receivable_amount = round_cash_amount(sum(calculate_remaining_repayment_amount(loan) for loan in issued_loans))
    today_new_receivable_amount = round_cash_amount(sum(calculate_total_repayment_amount(loan) for loan in today_issued_loans))

    overdue_outstanding_amount = round_cash_amount(sum(calculate_remaining_repayment_amount(loan) for loan in overdue_loans))
    yesterday_overdue_outstanding_amount = round_cash_amount(
        sum(
            calculate_remaining_repayment_amount(loan)
            for loan in overdue_loans
            if loan.due_date and loan.due_date < today_start
        )
    )

    pending_normal_outstanding_amount = round_cash_amount(
        sum(calculate_remaining_repayment_amount(loan) for loan in normal_outstanding_loans)
    )
    other_fee_amount = round_cash_amount(sum(float(loan.other_fee_amount or 0) for loan in issued_loans))

    cards = [
        {
            "key": "registered_users",
            "title": "注册数",
            "value": int(total_users),
            "value_type": "count",
            "sub_label": "今日注册数",
            "sub_value": int(today_new_users),
        },
        {
            "key": "order_count",
            "title": "下单数",
            "value": int(len(order_loans)),
            "value_type": "count",
            "sub_label": "今日下单数",
            "sub_value": int(today_order_count),
        },
        {
            "key": "ecard_issued_amount",
            "title": "E卡发放总额",
            "value": total_ecard_issued,
            "value_type": "currency",
            "sub_label": "今日E卡发放总额",
            "sub_value": today_ecard_issued,
        },
        {
            "key": "rights_cost",
            "title": "权益成本",
            "value": total_rights_cost,
            "value_type": "currency",
            "sub_label": "今日权益成本",
            "sub_value": today_rights_cost,
        },
        {
            "key": "overdue_amount",
            "title": "逾期总额",
            "value": overdue_outstanding_amount,
            "value_type": "currency",
            "sub_label": "昨日累计逾期金额",
            "sub_value": yesterday_overdue_outstanding_amount,
        },
        {
            "key": "issued_order_count",
            "title": "发卡单数",
            "value": int(len(issued_loans)),
            "value_type": "count",
            "sub_label": "今日发卡单数",
            "sub_value": int(len(today_issued_loans)),
        },
        {
            "key": "order_total_amount",
            "title": "订单总额",
            "value": order_total_amount,
            "value_type": "currency",
            "sub_label": "今日订单总额",
            "sub_value": today_order_total_amount,
        },
        {
            "key": "received_amount",
            "title": "收款金额",
            "value": round_cash_amount(sum(float(loan.repaid_amount or 0) for loan in issued_loans)),
            "value_type": "currency",
            "sub_label": "今日收款金额",
            "sub_value": round_cash_amount(today_received_amount),
        },
        {
            "key": "other_fee_amount",
            "title": "其他费用",
            "value": other_fee_amount,
            "value_type": "currency",
            "sub_label": "今日其他费用",
            "sub_value": round_cash_amount(today_other_fee_amount),
        },
        {
            "key": "receivable_amount",
            "title": "应收金额",
            "value": receivable_amount,
            "value_type": "currency",
            "sub_label": "今日新增应收金额",
            "sub_value": today_new_receivable_amount,
        },
        {
            "key": "pending_non_overdue_amount",
            "title": "待回收未逾期总资金",
            "value": pending_normal_outstanding_amount,
            "value_type": "currency",
            "sub_label": "--",
            "sub_value": 0,
        },
    ]
    charts = build_recent_insight_charts(issued_loans, today_start)

    return {
        "total_projects": 0,
        "total_borrowers": int(len({loan.user_id for loan in issued_loans})),
        "total_loans": int(len(issued_loans)),
        "total_payment_amount": 0,
        "total_receipt_amount": round_cash_amount(sum(float(loan.repaid_amount or 0) for loan in issued_loans)),
        "total_other_fee_amount": other_fee_amount,
        "total_net_amount": round_cash_amount(
            sum(float(loan.repaid_amount or 0) for loan in issued_loans)
            + other_fee_amount
            - total_ecard_issued
            - total_rights_cost
        ),
        "notes": [
            "服务费等科目的“付款”口径为减免/退费金额，不是用户额外打款。",
            "其他费用为账单外额外收取金额，不参与剩余待还计算。",
            "权益成本按每份已发出权益的权益定价 × 4% 计算。",
            "应收金额为截至当前未回收的应收总额（剩余待还）。",
        ],
        "cards": cards,
        "charts": charts,
        "items": [],
    }

async def _register_user(
    db: AsyncSession,
    current_admin: Admin,
    req: RegisterUserRequest,
):
    """后台新增前端用户。

    :param db: 异步数据库会话
    :param current_admin: 当前登录管理员
    :param req: 新增用户请求体
    :return: 新增结果
    """
    ensure_admin_page_permission(current_admin, "users")
    channel = None
    if req.source_channel_id is not None:
        channel = (
            await db.execute(
                select(Channel).where(
                    Channel.id == req.source_channel_id,
                    Channel.status == "ACTIVE",
                )
            )
        ).scalar_one_or_none()
        if channel is None:
            raise HTTPException(status_code=400, detail="来源渠道不存在或未启用")

    existed = (await db.execute(select(User).where(User.phone == req.phone))).scalar_one_or_none()
    if existed is not None:
        raise HTTPException(status_code=400, detail="手机号已存在")

    now = datetime.now()
    user = User(
        phone=req.phone,
        password_hash=get_password_hash(req.password),
        face_auth_status="PENDING",
        source_channel_id=channel.id if channel else None,
        channel_bound_at=now if channel else None,
        last_channel_visit_at=now if channel else None,
    )
    db.add(user)
    await db.flush()

    # 审计记录用于追踪管理员新增用户来源，避免后续归因排查困难。
    await log_user_event_async(
        db,
        user=user,
        loan=None,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_REGISTER_USER",
        title="后台新增用户",
        detail=(
            f"后台新增用户，来源渠道：{channel.sales_name}（{channel.channel_name}）。"
            if channel
            else "后台新增用户，来源渠道：未指定。"
        ),
    )
    await db.commit()
    return {
        "msg": "新增用户成功",
        "user_id": user.id,
        "phone": user.phone,
    }

async def _reset_user_password(
    db: AsyncSession,
    current_admin: Admin,
    user_id: int,
    req: ResetUserPasswordRequest,
):
    """后台重置指定用户密码。

    :param db: 异步数据库会话
    :param current_admin: 当前登录管理员
    :param user_id: 用户ID
    :param req: 重置密码请求体
    :return: 重置结果
    """
    ensure_admin_page_permission(current_admin, "users")
    if _is_business_consultant(current_admin):
        raise HTTPException(status_code=403, detail="业务顾问无权重置密码")
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="用户不存在")

    user.password_hash = get_password_hash(req.password)
    # 后台重置密码后立即解除该手机号冻结，避免用户已改密但仍被历史失败次数阻断。
    from app.api.endpoints import auth as auth_endpoints
    await auth_endpoints.password_login_guard.on_success(user.phone)
    await log_user_event_async(
        db,
        user=user,
        loan=None,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_RESET_PASSWORD",
        title="后台重置密码",
        detail="管理员已重置用户登录密码。",
    )
    await db.commit()
    return {"msg": "重置密码成功"}


async def _change_admin_password(
    db: AsyncSession,
    current_admin: Admin,
    req: AdminChangePasswordRequest,
):
    """当前登录后台用户修改自身密码。

    :param db: 异步数据库会话
    :param current_admin: 当前登录管理员
    :param req: 修改密码请求体
    :return: 修改结果
    """
    if req.new_password != req.confirm_password:
        raise HTTPException(status_code=400, detail="两次输入的新密码不一致")
    if not verify_password(req.old_password, current_admin.password_hash):
        raise HTTPException(status_code=400, detail="原密码错误")
    if req.old_password == req.new_password:
        raise HTTPException(status_code=400, detail="新密码不能与原密码一致")

    current_admin.password_hash = get_password_hash(req.new_password)
    await db.commit()
    return {"msg": "修改密码成功"}

async def _get_loan_ledger(db: AsyncSession, current_admin: Admin, loan_id: int):
    loan = (
        await db.execute(
            select(Loan)
            .options(
                joinedload(Loan.installments),
                joinedload(Loan.transactions),
                joinedload(Loan.review_admin),
                joinedload(Loan.collection_admin),
            )
            .where(Loan.id == loan_id)
        )
    ).unique().scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")

    if loan.status == "OVERDUE":
        ensure_any_admin_page_permission(current_admin, ("repayments", "collections"))
    else:
        permission_key = resolve_loan_scope_permission(None, None, loan.status)
        if permission_key:
            ensure_admin_page_permission(current_admin, permission_key)
        else:
            ensure_any_admin_page_permission(current_admin, LOAN_PAGE_PERMISSION_KEYS)
    ensure_stage_access_for_admin(current_admin, loan)

    ledger = get_loan_ledger_snapshot(loan)
    return {
        "loan_id": loan.id,
        "loan_status": loan.status,
        "installments": ledger["installments"],
        "transactions": [serialize_transaction(item) for item in loan.transactions],
        "fund_flow_summary": ledger["summary"],
    }

async def _get_user_detail(db: AsyncSession, current_admin: Admin, user_id: int):
    ensure_admin_page_permission(current_admin, "users")
    user = (
        await db.execute(
            select(User)
            .options(
                joinedload(User.loans),
                joinedload(User.source_channel),
            )
            .where(User.id == user_id)
        )
    ).unique().scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if _is_business_consultant(current_admin):
        bound_channel = getattr(user, "source_channel", None)
        if not bound_channel or int(bound_channel.admin_user_id or 0) != int(current_admin.id):
            raise HTTPException(status_code=403, detail="无权查看该用户档案")
    events = (
        await db.execute(
            select(UserEvent)
            .where(UserEvent.user_id == user.id)
            .order_by(UserEvent.created_at.desc())
        )
    ).scalars().all()
    payload = serialize_user_detail(user, events=events)
    payload["can_unlock_location_risk"] = admin_has_permission(current_admin, "user-location-risk-unlock")
    return payload


async def _unlock_user_location_risk(
    db: AsyncSession,
    current_admin: Admin,
    user_id: int,
):
    """解除指定用户的位置风控锁定，但保留原有位置记录。

    :param db: 异步数据库会话
    :param current_admin: 当前登录管理员
    :param user_id: 用户ID
    :return: 处理结果
    """
    ensure_admin_page_permission(current_admin, "users")
    ensure_admin_permission(current_admin, "user-location-risk-unlock")
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 只解除当前锁定状态，保留历史GPS/IP轨迹，方便后续复核。
    user.location_risk_blocked = False
    user.location_risk_reason = None
    user.location_risk_at = None

    await log_user_event_async(
        db,
        user=user,
        loan=None,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_LOCATION_RISK_UNLOCK",
        title="后台解除位置风控",
        detail="管理员手动解除位置风控锁定，历史定位记录保留。",
    )
    await db.commit()
    return {"msg": "位置风控已解除"}

async def _get_risk_report(db: AsyncSession, req: AdminRiskReportRequest, current_admin: Admin):
    """兼容旧风控查询入口，统一返回 GalaCredit 风险报告。

    :param db: 异步数据库会话
    :param req: 风控报告请求
    :param current_admin: 当前后台用户
    :return: GalaCredit 风险报告
    """
    return await _get_composite_risk_report(db, req, current_admin)

async def _get_composite_risk_report(db: AsyncSession, req: AdminRiskReportRequest, current_admin: Admin):
    """查询 GalaCredit 风险报告。

    :param db: 异步数据库会话
    :param req: 风控报告请求
    :param current_admin: 当前后台用户
    :return: GalaCredit 风险报告
    """
    ensure_any_admin_page_permission(current_admin, ("users", "applications", "disbursements", "repayments", "collections", "financials"))
    user = await get_user_for_risk_report_async(db, req.user_id)
    report = await get_or_create_composite_risk_report_async(db, user=user)
    latest_loan = await get_latest_loan_async(db, user.id)
    if latest_loan:
        latest_loan.risk_report_checked_at = datetime.now()
        latest_loan.risk_report_checked_by = current_admin.username
    await log_user_event_async(
        db,
        user=user,
        loan=latest_loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_COMPOSITE_RISK_REPORT",
        title="查询 GalaCredit 风险报告",
        detail="后台发起 GalaCredit 风险报告查询",
    )
    await db.commit()
    await db.refresh(report)
    return serialize_composite_risk_report(report)

def _normalize_single_risk_value(value: Optional[str]) -> str:
    """规范化单查输入值。

    :param value: 原始输入值
    :return: 去除首尾空格后的字符串
    """
    return str(value or "").strip()

async def _resolve_single_risk_subject(
    db: AsyncSession,
    *,
    name: str,
    id_card: str,
    phone: str,
) -> tuple[str, str, str, Optional[int]]:
    """根据可选三要素解析最终查询对象。

    :param db: 异步数据库会话
    :param name: 姓名
    :param id_card: 身份证号
    :param phone: 手机号
    :return: 姓名、身份证、手机号、用户ID
    """
    if not any([name, id_card, phone]):
        raise HTTPException(status_code=400, detail="请至少填写姓名、身份证号或手机号中的一项")

    filters = []
    if name:
        filters.append(User.name == name)
    if id_card:
        filters.append(User.id_card_num == id_card)
    if phone:
        filters.append(User.phone == phone)
    users = (await db.execute(select(User).where(*filters).order_by(User.id.desc()).limit(2))).scalars().all()
    if len(users) > 1:
        raise HTTPException(status_code=400, detail="匹配到多个客户，请补充更完整的三要素后再查询")
    if len(users) == 1:
        user = users[0]
        resolved_name = name or _normalize_single_risk_value(user.name)
        resolved_id_card = id_card or _normalize_single_risk_value(user.id_card_num)
        resolved_phone = phone or _normalize_single_risk_value(user.phone)
        if not all([resolved_name, resolved_id_card, resolved_phone]):
            raise HTTPException(status_code=400, detail="匹配客户实名信息不完整，请补齐三要素后再查询")
        return resolved_name, resolved_id_card, resolved_phone, user.id

    if not all([name, id_card, phone]):
        raise HTTPException(status_code=400, detail="系统内未匹配到唯一客户，请补齐姓名、身份证号和手机号后再单查")
    return name, id_card, phone, None

async def _query_single_risk_report(
    db: AsyncSession,
    req: AdminRiskSingleReportRequest,
    current_admin: Admin,
):
    """执行风控报告单查。

    :param db: 异步数据库会话
    :param req: 单查请求
    :param current_admin: 当前后台用户
    :return: GalaCredit 风险报告
    """
    ensure_admin_page_permission(current_admin, "risk-single-query")
    name = _normalize_single_risk_value(req.name)
    id_card = _normalize_single_risk_value(req.id_card)
    phone = _normalize_single_risk_value(req.phone)
    resolved_name, resolved_id_card, resolved_phone, user_id = await _resolve_single_risk_subject(
        db,
        name=name,
        id_card=id_card,
        phone=phone,
    )
    if user_id:
        user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")
        # 已有客户必须复用当前系统用户报告链路，确保弹窗内容与用户详情中完全一致。
        report = await get_or_create_composite_risk_report_async(db, user=user)
    else:
        report = await get_or_create_standalone_composite_risk_report_async(
            db,
            name=resolved_name,
            id_card=resolved_id_card,
            phone=resolved_phone,
            user_id=user_id,
        )
    await db.commit()
    await db.refresh(report)
    return serialize_composite_risk_report(report)

async def _get_single_risk_report_history(
    db: AsyncSession,
    current_admin: Admin,
    keyword: Optional[str],
    skip: int,
    limit: int,
):
    """获取风控报告查询历史。

    :param db: 异步数据库会话
    :param current_admin: 当前后台用户
    :param keyword: 客户三要素关键词
    :param skip: 跳过条数
    :param limit: 返回条数
    :return: 分页历史清单
    """
    ensure_admin_page_permission(current_admin, "risk-single-query")
    limit = min(max(limit, 1), 100)
    skip = max(skip, 0)
    stmt = select(RiskCompositeReport)
    count_stmt = select(func.count(RiskCompositeReport.id))
    if keyword:
        pattern = f"%{keyword.strip()}%"
        keyword_filter = or_(
            RiskCompositeReport.name.like(pattern),
            RiskCompositeReport.id_card.like(pattern),
            RiskCompositeReport.phone.like(pattern),
        )
        stmt = stmt.where(keyword_filter)
        count_stmt = count_stmt.where(keyword_filter)
    total = (await db.execute(count_stmt)).scalar_one()
    rows = (
        await db.execute(
            stmt.order_by(RiskCompositeReport.query_time.desc(), RiskCompositeReport.id.desc())
            .offset(skip)
            .limit(limit)
        )
    ).scalars().all()
    return {
        "total": total,
        "page": skip // limit + 1,
        "size": limit,
        "items": [
            {
                "id": item.id,
                "user_id": item.user_id,
                "name": item.name,
                "id_card": item.id_card,
                "phone": item.phone,
                "query_time": item.query_time,
                "created_at": item.created_at,
            }
            for item in rows
        ],
    }

async def _get_single_risk_report_detail(
    db: AsyncSession,
    current_admin: Admin,
    report_id: int,
):
    """根据历史记录ID获取风控报告详情。

    :param db: 异步数据库会话
    :param current_admin: 当前后台用户
    :param report_id: 报告ID
    :return: GalaCredit 风险报告
    """
    ensure_admin_page_permission(current_admin, "risk-single-query")
    report = (
        await db.execute(select(RiskCompositeReport).where(RiskCompositeReport.id == report_id))
    ).scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="风控报告不存在")
    return serialize_composite_risk_report(report)

async def _get_channels(db: AsyncSession, current_admin: Admin, keyword: Optional[str], status: Optional[str], skip: int, limit: int):
    ensure_admin_page_permission(current_admin, "channels")
    limit = min(max(limit, 1), 100)
    stmt = select(Channel).options(joinedload(Channel.users).joinedload(User.loans))
    if keyword:
        pattern = f"%{keyword.strip().lower()}%"
        stmt = stmt.where(
            or_(
                Channel.channel_name.like(pattern),
                Channel.sales_name.like(f"%{keyword.strip()}%"),
            )
        )
    if status and status != "ALL":
        stmt = stmt.where(Channel.status == normalize_channel_status(status))
    matched_channels = (await db.execute(stmt.order_by(Channel.created_at.desc()))).unique().scalars().all()
    advisor_ids = {int(item.admin_user_id) for item in matched_channels if item.admin_user_id}
    advisor_map = {}
    if advisor_ids:
        advisors = (await db.execute(select(Admin).where(Admin.id.in_(advisor_ids)))).scalars().all()
        advisor_map = {int(item.id): item for item in advisors}
    channel_items = [serialize_channel(channel, advisor_map.get(int(channel.admin_user_id or 0))) for channel in matched_channels]
    return {
        "total": len(channel_items),
        "page": skip // limit + 1,
        "size": limit,
        "channel_link_prefix": settings.CHANNEL_LINK_PREFIX,
        "summary": build_channel_summary(channel_items),
        "items": channel_items[skip : skip + limit],
    }


async def _get_exclusive_links(db: AsyncSession, current_admin: Admin):
    """获取当前业务顾问可见的专属链接列表。

    :param db: 异步数据库会话
    :param current_admin: 当前登录管理员
    :return: 专属链接列表（最多 50 条，启用中优先）
    """
    ensure_admin_page_permission(current_admin, "exclusive-links")
    stmt = (
        select(Channel)
        .options(joinedload(Channel.users).joinedload(User.loans))
        .where(Channel.admin_user_id == current_admin.id)
        .order_by(
            case((Channel.status == "ACTIVE", 0), else_=1),
            Channel.created_at.desc(),
        )
        .limit(50)
    )
    channels = (await db.execute(stmt)).unique().scalars().all()
    channel_items = [serialize_channel(channel, current_admin) for channel in channels]
    return {
        "channel_link_prefix": settings.CHANNEL_LINK_PREFIX,
        "items": channel_items,
    }


async def _get_user_source_channels(db: AsyncSession, current_admin: Admin, keyword: Optional[str], limit: int):
    """获取“新增用户”可选来源渠道。

    :param db: 异步数据库会话
    :param current_admin: 当前登录管理员
    :param keyword: 渠道关键词
    :param limit: 返回数量限制
    :return: 渠道列表
    """
    ensure_admin_page_permission(current_admin, "users")
    limit = min(max(int(limit or 20), 1), 50)
    stmt = select(Channel).where(Channel.status == "ACTIVE")
    if keyword:
        pattern = f"%{keyword.strip()}%"
        stmt = stmt.where(or_(Channel.channel_name.like(pattern), Channel.sales_name.like(pattern)))
    if _is_business_consultant(current_admin):
        stmt = stmt.where(Channel.admin_user_id == current_admin.id)
    channels = (await db.execute(stmt.order_by(Channel.id.asc()).limit(limit))).scalars().all()
    return [serialize_channel_landing(item) for item in channels]

async def _create_channel(db: AsyncSession, current_admin: Admin, req: ChannelCreateRequest):
    ensure_admin_page_permission(current_admin, "channels")
    channel_name = normalize_channel_name(req.channel_name)
    sales_name = req.sales_name.strip()
    if not sales_name:
        raise HTTPException(status_code=400, detail="请填写业务员姓名")
    exists = (await db.execute(select(Channel).where(Channel.channel_name == channel_name))).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=400, detail="渠道名称已存在")
    invite_code = (req.invite_code or "").strip().lower()
    if not invite_code:
        invite_code = await _generate_unique_channel_invite_code(db, 16)
    else:
        invite_code_exists = (await db.execute(select(Channel).where(Channel.invite_code == invite_code))).scalar_one_or_none()
        if invite_code_exists:
            raise HTTPException(status_code=400, detail="渠道邀请码已存在")
    advisor = await _resolve_business_advisor_by_id(db, req.admin_user_id)
    channel = Channel(
        channel_name=channel_name,
        invite_code=invite_code,
        sales_name=sales_name,
        status=normalize_channel_status(req.status),
        disbursement_mode=normalize_channel_disbursement_mode(req.disbursement_mode),
        review_mode=normalize_channel_review_mode(req.review_mode),
        note=(req.note or "").strip() or None,
        admin_user_id=advisor.id,
    )
    db.add(channel)
    await db.commit()
    await db.refresh(channel)
    db.add(ConfigChangeHistory(object_type="CHANNEL", object_id=channel.id, action="CREATE", version_no=1, snapshot_json=json.dumps(serialize_channel(channel), default=str, ensure_ascii=False), operator_name=current_admin.username))
    await db.commit()
    return serialize_channel(channel, advisor)

async def _update_channel(db: AsyncSession, current_admin: Admin, channel_id: int, req: ChannelUpdateRequest):
    ensure_admin_page_permission(current_admin, "channels")
    channel = (
        await db.execute(
            select(Channel)
            .options(joinedload(Channel.users).joinedload(User.loans))
            .where(Channel.id == channel_id)
        )
    ).unique().scalar_one_or_none()
    if not channel:
        raise HTTPException(status_code=404, detail="渠道不存在")
    payload = req.model_dump(exclude_unset=True)
    if "sales_name" in payload and payload["sales_name"] is not None:
        sales_name = payload["sales_name"].strip()
        if not sales_name:
            raise HTTPException(status_code=400, detail="请填写业务员姓名")
        channel.sales_name = sales_name
    if "status" in payload and payload["status"] is not None:
        channel.status = normalize_channel_status(payload["status"])
    if "disbursement_mode" in payload and payload["disbursement_mode"] is not None:
        channel.disbursement_mode = normalize_channel_disbursement_mode(payload["disbursement_mode"])
    if "review_mode" in payload and payload["review_mode"] is not None:
        channel.review_mode = normalize_channel_review_mode(payload["review_mode"])
    if "note" in payload:
        channel.note = (payload["note"] or "").strip() or None
    if "admin_user_id" in payload and payload["admin_user_id"] is not None:
        advisor = await _resolve_business_advisor_by_id(db, int(payload["admin_user_id"]))
        channel.admin_user_id = advisor.id
    if not (channel.invite_code or "").strip():
        # 兼容历史空邀请码数据：编辑保存时自动补齐，避免专属链接不可用
        channel.invite_code = await _generate_unique_channel_invite_code(db, 16)
    await db.commit()
    await db.refresh(channel)
    version = int((await db.scalar(select(func.coalesce(func.max(ConfigChangeHistory.version_no), 0)).where(ConfigChangeHistory.object_type == "CHANNEL", ConfigChangeHistory.object_id == channel.id))) or 0) + 1
    db.add(ConfigChangeHistory(object_type="CHANNEL", object_id=channel.id, action="UPDATE", version_no=version, snapshot_json=json.dumps(serialize_channel(channel), default=str, ensure_ascii=False), operator_name=current_admin.username))
    await db.commit()
    advisor = None
    if channel.admin_user_id:
        advisor = (await db.execute(select(Admin).where(Admin.id == channel.admin_user_id))).scalar_one_or_none()
    return serialize_channel(channel, advisor)


async def _get_business_advisors(db: AsyncSession, current_admin: Admin, keyword: Optional[str], limit: int):
    ensure_admin_page_permission(current_admin, "channels")
    limit = min(max(limit, 1), 50)
    stmt = select(Admin)
    if keyword:
        text = keyword.strip()
        if text:
            if text.isdigit():
                stmt = stmt.where(or_(Admin.id == int(text), Admin.username.like(f"%{text}%")))
            else:
                stmt = stmt.where(Admin.username.like(f"%{text}%"))
    admins = (await db.execute(stmt.order_by(Admin.id.desc()).limit(limit * 3))).scalars().all()
    items: list[BusinessAdvisorItemResponse] = []
    for admin in admins:
        if _is_business_consultant(admin):
            items.append(BusinessAdvisorItemResponse(id=admin.id, username=admin.username))
            if len(items) >= limit:
                break
    return items

async def _get_products(db: AsyncSession, current_admin: Admin, keyword: Optional[str], is_active: Optional[bool], skip: int, limit: int):
    ensure_admin_page_permission(current_admin, "products")
    limit = min(max(limit, 1), 100)
    stmt = select(Product)
    if keyword:
        stmt = stmt.where(Product.name.like(f"%{keyword.strip()}%"))
    if is_active is not None:
        stmt = stmt.where(Product.is_active.is_(is_active))
    total = (await db.scalar(select(func.count()).select_from(stmt.subquery()))) or 0
    items = (
        await db.execute(stmt.order_by(Product.updated_at.desc(), Product.id.desc()).offset(skip).limit(limit))
    ).scalars().all()
    return {
        "total": total,
        "page": skip // limit + 1,
        "size": limit,
        "items": [serialize_product(item) for item in items],
    }


async def _get_disbursement_failures(
    db: AsyncSession,
    current_admin: Admin,
    keyword: Optional[str],
    skip: int,
    limit: int,
):
    """查询每笔订单最近一次失败的 MoMo 放款记录。

    :param db: 异步数据库会话
    :param current_admin: 当前管理员
    :param keyword: 手机号、姓名或身份证号筛选词
    :param skip: 分页偏移量
    :param limit: 分页数量
    :return: 放款失败客户分页数据
    """
    ensure_admin_page_permission(current_admin, "disbursement-failures")
    limit = min(max(int(limit or 20), 1), 100)
    latest_failure_ids = (
        select(func.max(MomoTransaction.id).label("transaction_id"))
        .where(
            MomoTransaction.transaction_type == "DISBURSEMENT",
            MomoTransaction.status == "FAILED",
        )
        .group_by(MomoTransaction.loan_id)
        .subquery()
    )
    stmt = (
        select(MomoTransaction, Loan, User)
        .join(Loan, Loan.id == MomoTransaction.loan_id)
        .join(User, User.id == MomoTransaction.user_id)
        .where(MomoTransaction.id.in_(select(latest_failure_ids.c.transaction_id)))
    )
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        stmt = stmt.where(or_(User.phone.like(pattern), User.name.like(pattern), User.id_card_num.like(pattern)))

    base_query = stmt.subquery()
    total = int(await db.scalar(select(func.count()).select_from(base_query)) or 0)
    amount = float(await db.scalar(select(func.coalesce(func.sum(base_query.c.amount), 0))) or 0)
    rows = (await db.execute(stmt.order_by(MomoTransaction.id.desc()).offset(skip).limit(limit))).all()
    return {
        "total": total,
        "page": skip // limit + 1,
        "size": limit,
        "summary": {"failed_count": total, "failed_amount": round_money(amount)},
        "items": [
            {
                "transaction_id": transaction.id,
                "loan_id": loan.id,
                "user_id": user.id,
                "user_name": user.name,
                "user_phone": user.phone,
                "id_card_num": mask_secret(user.id_card_num, left=4, right=2),
                "loan_status": loan.status,
                "product_name": loan.product_name,
                "nominal_loan_amount": round_money(loan.nominal_loan_amount or loan.credit_limit),
                "actual_disbursement_amount": round_money(loan.actual_disbursement_amount),
                "amount": round_money(transaction.amount),
                "provider": transaction.provider,
                "provider_reference": transaction.provider_reference,
                "failure_message": transaction.failure_message or "MoMo disbursement failed",
                "requested_at": transaction.requested_at,
                "completed_at": transaction.completed_at,
                "retryable": loan.status == "WITHDRAWING" and not user.blacklist_hit,
                "user_blacklist_hit": bool(user.blacklist_hit),
            }
            for transaction, loan, user in rows
        ],
    }

async def _create_product(db: AsyncSession, current_admin: Admin, req: ProductCreateRequest):
    ensure_admin_page_permission(current_admin, "products")
    if req.product_type == "ECARD_RIGHTS" and round_money(req.ecard_face_value) <= 0:
        raise HTTPException(status_code=400, detail="E卡+权益商品必须填写E卡面值")
    if req.product_type == "RIGHTS_ONLY" and round_money(req.rights_price) <= 0:
        raise HTTPException(status_code=400, detail="纯权益包必须填写权益金额")
    payment_amount = round_money(req.nominal_loan_amount or req.payment_amount or resolve_product_payment_amount(req.ecard_face_value, req.rights_price))
    if req.product_type == "CASH_LOAN":
        fee_components = normalize_cash_loan_fee_components(req.fee_components)
        await validate_cash_loan_compliance_async(
            db,
            nominal_amount=payment_amount,
            upfront_fee_rate=req.upfront_fee_rate,
            repayment_due_day=req.repayment_due_day,
            daily_overdue_fee=req.daily_overdue_fee,
            term_days=req.term_days,
            installment_count=req.installment_count,
            fee_components=fee_components,
        )
    else:
        fee_components = req.fee_components
    product = Product(
        name=req.name.strip(),
        ecard_face_value=round_money(req.ecard_face_value),
        rights_price=round_money(req.rights_price),
        rights_title=req.rights_title.strip(),
        rights_desc=(req.rights_desc or "").strip() or None,
        rights_detail_json=json.dumps(req.rights_detail, ensure_ascii=False) if req.rights_detail else None,
        term_days=req.term_days,
        payment_amount=payment_amount,
        nominal_loan_amount=payment_amount,
        upfront_fee_rate=req.upfront_fee_rate,
        fee_components_json=json.dumps(fee_components, ensure_ascii=False) if fee_components else None,
        interest_start_day=req.interest_start_day,
        repayment_due_day=req.repayment_due_day,
        installment_count=req.installment_count,
        installment_ratios_json=json.dumps(req.installment_ratios, ensure_ascii=False) if req.installment_ratios else None,
        daily_overdue_fee=req.daily_overdue_fee,
        borrower_type=req.borrower_type,
        product_type=req.product_type,
        is_active=req.is_active,
    )
    db.add(product)
    await db.commit()
    await db.refresh(product)
    db.add(ConfigChangeHistory(object_type="PRODUCT", object_id=product.id, action="CREATE", version_no=1, snapshot_json=json.dumps(serialize_product(product), default=str, ensure_ascii=False), operator_name=current_admin.username))
    await db.commit()
    return serialize_product(product)

async def _update_product(db: AsyncSession, current_admin: Admin, product_id: int, req: ProductUpdateRequest):
    ensure_admin_page_permission(current_admin, "products")
    product = (await db.execute(select(Product).where(Product.id == product_id))).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    payload = req.model_dump(exclude_unset=True)
    if "name" in payload and payload["name"] is not None:
        product.name = payload["name"].strip()
    if "ecard_face_value" in payload and payload["ecard_face_value"] is not None:
        product.ecard_face_value = round_money(payload["ecard_face_value"])
    if "rights_price" in payload and payload["rights_price"] is not None:
        product.rights_price = round_money(payload["rights_price"])
    if "rights_title" in payload and payload["rights_title"] is not None:
        product.rights_title = payload["rights_title"].strip()
    if "rights_desc" in payload:
        product.rights_desc = (payload["rights_desc"] or "").strip() or None
    if "rights_detail" in payload:
        product.rights_detail_json = json.dumps(payload["rights_detail"], ensure_ascii=False) if payload["rights_detail"] else None
    if "term_days" in payload and payload["term_days"] is not None:
        product.term_days = payload["term_days"]
    if "is_active" in payload and payload["is_active"] is not None:
        product.is_active = bool(payload["is_active"])
    if "payment_amount" in payload and payload["payment_amount"] is not None:
        product.payment_amount = round_money(payload["payment_amount"])
        product.nominal_loan_amount = product.payment_amount
    if "nominal_loan_amount" in payload and payload["nominal_loan_amount"] is not None:
        product.nominal_loan_amount = round_money(payload["nominal_loan_amount"])
        product.payment_amount = product.nominal_loan_amount
    for field in ("upfront_fee_rate", "interest_start_day", "repayment_due_day", "installment_count", "daily_overdue_fee"):
        if field in payload and payload[field] is not None:
            setattr(product, field, payload[field])
    if "fee_components" in payload:
        product.fee_components_json = json.dumps(payload["fee_components"], ensure_ascii=False) if payload["fee_components"] else None
    if "installment_ratios" in payload:
        product.installment_ratios_json = json.dumps(payload["installment_ratios"], ensure_ascii=False) if payload["installment_ratios"] else None
    elif "ecard_face_value" in payload or "rights_price" in payload:
        product.payment_amount = resolve_product_payment_amount(product.ecard_face_value, product.rights_price)
    if "product_type" in payload and payload["product_type"] is not None:
        product.product_type = payload["product_type"]
    if "borrower_type" in payload and payload["borrower_type"] is not None:
        product.borrower_type = payload["borrower_type"]
    if product.product_type == "ECARD_RIGHTS" and round_money(product.ecard_face_value) <= 0:
        raise HTTPException(status_code=400, detail="E卡+权益商品必须填写E卡面值")
    if product.product_type == "RIGHTS_ONLY" and round_money(product.rights_price) <= 0:
        raise HTTPException(status_code=400, detail="纯权益包必须填写权益金额")
    if product.product_type == "CASH_LOAN":
        fee_components = normalize_cash_loan_fee_components(payload.get("fee_components") if "fee_components" in payload else json.loads(product.fee_components_json or "{}"))
        await validate_cash_loan_compliance_async(
            db,
            nominal_amount=product.nominal_loan_amount or product.payment_amount,
            upfront_fee_rate=product.upfront_fee_rate,
            repayment_due_day=product.repayment_due_day,
            daily_overdue_fee=product.daily_overdue_fee,
            term_days=product.term_days,
            installment_count=product.installment_count,
            fee_components=fee_components,
        )
        product.fee_components_json = json.dumps(fee_components, ensure_ascii=False)
    await db.commit()
    await db.refresh(product)
    version = int((await db.scalar(select(func.coalesce(func.max(ConfigChangeHistory.version_no), 0)).where(ConfigChangeHistory.object_type == "PRODUCT", ConfigChangeHistory.object_id == product.id))) or 0) + 1
    db.add(ConfigChangeHistory(object_type="PRODUCT", object_id=product.id, action="UPDATE", version_no=version, snapshot_json=json.dumps(serialize_product(product), default=str, ensure_ascii=False), operator_name=current_admin.username))
    await db.commit()
    return serialize_product(product)

async def _get_ecard_pool(db: AsyncSession, current_admin: Admin, keyword: Optional[str], status: Optional[str], face_value: Optional[float], skip: int, limit: int):
    ensure_admin_page_permission(current_admin, "ecard-pool")
    limit = min(max(limit, 1), 100)
    stmt = select(EcardPool)
    if keyword:
        stmt = stmt.where(EcardPool.account.like(f"%{keyword.strip()}%"))
    if status and status != "ALL":
        upper_status = status.upper()
        if upper_status not in ECARD_POOL_STATUSES:
            raise HTTPException(status_code=400, detail="卡池状态非法")
        stmt = stmt.where(EcardPool.status == upper_status)
    if face_value is not None:
        stmt = stmt.where(EcardPool.face_value == round_money(face_value))
    total = (await db.scalar(select(func.count()).select_from(stmt.subquery()))) or 0
    items = (
        await db.execute(stmt.order_by(EcardPool.expires_at.asc(), EcardPool.id.desc()).offset(skip).limit(limit))
    ).scalars().all()
    await _attach_ecard_pool_display_fields(db, items)
    return {
        "total": total,
        "page": skip // limit + 1,
        "size": limit,
        "stats": await _build_ecard_pool_stats(db),
        "items": [serialize_ecard_pool_item(item) for item in items],
    }

async def _create_ecard_pool_item(db: AsyncSession, current_admin: Admin, req: EcardPoolCreateRequest):
    ensure_admin_page_permission(current_admin, "ecard-pool")
    account = req.account.strip()
    if (await db.execute(select(EcardPool).where(EcardPool.account == account))).scalar_one_or_none():
        raise HTTPException(status_code=400, detail="卡号已存在")
    item = EcardPool(
        account=account,
        password=req.password.strip(),
        face_value=round_money(req.face_value),
        expires_at=req.expires_at,
        status="AVAILABLE",
        note=(req.note or "").strip() or None,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return serialize_ecard_pool_item(item)

def _parse_upload_expiration(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError("有效期不能为空")
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            raise ValueError("有效期格式必须为 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS")
    raise ValueError("有效期格式错误")

def _excel_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()

def _load_excel_rows(upload_file: UploadFile):
    content = upload_file.file.read()
    upload_file.file.close()
    filename = upload_file.filename.lower()
    rows = []
    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            rows.append(list(row))
    elif filename.endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        for row_idx in range(1, sheet.nrows):
            row_values = []
            for col_idx in range(4):
                cell = sheet.cell(row_idx, col_idx)
                if col_idx == 3 and cell.ctype == xlrd.XL_CELL_DATE:
                    row_values.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
                else:
                    row_values.append(cell.value)
            rows.append(row_values)
    else:
        raise HTTPException(status_code=400, detail="仅支持 xls 或 xlsx 文件")
    return rows

async def _upload_ecard_pool_items(db: AsyncSession, file: UploadFile, current_admin: Admin):
    ensure_admin_page_permission(current_admin, "ecard-pool")
    rows = _load_excel_rows(file)
    if not rows:
        raise HTTPException(status_code=400, detail="上传文件内容不能为空")

    upload_accounts = [_excel_text(row[0]) for row in rows if row and _excel_text(row[0])]
    existing_accounts = set()
    if upload_accounts:
        existing_accounts = set(
            (await db.execute(select(EcardPool.account).where(EcardPool.account.in_(upload_accounts)))).scalars().all()
        )
    created = 0
    errors = []
    seen_accounts = set()

    for index, row in enumerate(rows, start=2):
        account = _excel_text(row[0]) if len(row) >= 1 else ""
        password = _excel_text(row[1]) if len(row) >= 2 else ""
        face_value = row[2] if len(row) >= 3 else None
        expires_at = row[3] if len(row) >= 4 else None

        if not account and not password and not face_value and not expires_at:
            continue
        if not account:
            errors.append({"row": index, "reason": "卡号不能为空"})
            continue
        if not password:
            errors.append({"row": index, "reason": "密码不能为空"})
            continue
        if account in seen_accounts:
            errors.append({"row": index, "reason": "文件中存在重复卡号"})
            continue
        if account in existing_accounts:
            errors.append({"row": index, "reason": "卡号已存在"})
            continue

        try:
            face_value = float(face_value)
            if face_value <= 0:
                raise ValueError
        except Exception:
            errors.append({"row": index, "reason": "面额必须为正数"})
            continue

        try:
            expires_at = _parse_upload_expiration(expires_at)
        except ValueError as exc:
            errors.append({"row": index, "reason": str(exc)})
            continue

        item = EcardPool(
            account=account,
            password=password,
            face_value=round_money(face_value),
            expires_at=expires_at,
            status="AVAILABLE",
        )
        db.add(item)
        seen_accounts.add(account)
        created += 1

    await db.commit()
    return {
        "created": created,
        "errors": errors,
    }

async def _update_ecard_pool_item(db: AsyncSession, current_admin: Admin, item_id: int, req: EcardPoolUpdateRequest):
    ensure_admin_page_permission(current_admin, "ecard-pool")
    item = (await db.execute(select(EcardPool).where(EcardPool.id == item_id))).scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="卡池记录不存在")

    payload = req.model_dump(exclude_unset=True)
    if "status" in payload and payload["status"] is not None:
        next_status = payload["status"].upper()
        if next_status not in ECARD_POOL_STATUSES:
            raise HTTPException(status_code=400, detail="卡池状态非法")
        if item.status == "ASSIGNED" and next_status != "ASSIGNED":
            raise HTTPException(status_code=400, detail="已发放卡不可手工变更状态")
        item.status = next_status
    if "note" in payload:
        item.note = (payload["note"] or "").strip() or None
    if "expires_at" in payload and payload["expires_at"] is not None:
        item.expires_at = payload["expires_at"]

    await db.commit()
    await db.refresh(item)
    return serialize_ecard_pool_item(item)

async def _review_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req: LoanReviewRequest):
    ensure_admin_page_permission(current_admin, "applications")
    loan = (
        await db.execute(
            select(Loan)
            .options(joinedload(Loan.owner), joinedload(Loan.review_admin))
            .where(Loan.id == loan_id)
        )
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status in {"WITHDRAWING", "DISBURSED", "OVERDUE", "SETTLED"}:
        raise HTTPException(status_code=400, detail="当前订单已进入发卡/付款流程，不能重新审批")

    await assign_review_admin_if_needed_async(db, loan)
    if not is_super_admin(current_admin):
        if int(loan.review_admin_id or 0) != int(current_admin.id):
            raise HTTPException(status_code=403, detail="仅可处理分配给你的审批订单")

    owner = loan.owner
    await refresh_user_blacklist_status(db, owner)
    if req.approved:
        if owner.blacklist_hit:
            raise HTTPException(status_code=400, detail="该用户命中黑名单，只能操作拒绝")
        if not loan.risk_report_checked_at:
            raise HTTPException(status_code=400, detail="审批通过前请先查看风控报告")
        if loan.risk_report_checked_at < datetime.now() - timedelta(days=14):
            raise HTTPException(status_code=400, detail="风控报告已超过14天，请重新查询后再审批通过")
        if loan.risk_report_checked_by and loan.risk_report_checked_by != current_admin.username:
            raise HTTPException(status_code=400, detail="当前审核员审批通过前请先查看风控报告")
        if req.credit_limit is None:
            raise HTTPException(status_code=400, detail="请填写信用额度")
        approved_credit_limit = round_money(req.credit_limit)
        term_days = ensure_valid_term_days(req.term_days or 7)
        discount_amount = round_money(req.approval_discount_amount or 0)
        if discount_amount > approved_credit_limit:
            raise HTTPException(status_code=400, detail="减免额度不能超过授信额度")

        loan.status = "APPROVED"
        loan.approved_credit_limit = approved_credit_limit
        loan.credit_limit = approved_credit_limit
        loan.approval_discount_amount = discount_amount
        loan.order_discount_amount = 0
        loan.fee_rate = DEFAULT_FEE_RATE
        loan.fee_amount = 0
        loan.term_days = term_days
        loan.product_term_days = term_days
        loan.review_note = (req.review_note or "后台已完成授信审批").strip()
        loan.approved_at = datetime.now()
        loan.due_date = None
        loan.disbursed_at = None
        loan.penalty_amount = 0
        loan.repaid_amount = 0
        loan.reduction_amount = 0
        loan.other_fee_amount = 0
        loan.paid_penalty_amount = 0
        loan.reduced_penalty_amount = 0
        loan.actual_repayment_date = None
        loan.reminder_count = 0
        loan.last_reminded_at = None
        loan.collection_count = 0
        loan.last_collection_at = None
        loan.collection_note = None
        loan.collection_admin_id = None
        loan.collection_transferred_at = None
        loan.repay_attempt_count = 0
        loan.product_id = None
        loan.product_name = None
        loan.rights_title = None
        loan.rights_desc = None
        loan.rights_contact_phone = None
        loan.rights_price = 0
        loan.ecard_face_value = 0
        loan.product_total_price = 0
        loan.ecard_account = None
        loan.ecard_password = None
        loan.ecard_expires_at = None
        owner.approved_limit = int(approved_credit_limit)
        _set_available_credit(owner, approved_credit_limit)
        owner.overdue_credit_locked = False

        detail = (
            f"审批通过；授信额度 {approved_credit_limit:.2f} 元；"
            f"期限 {term_days} 天；"
            f"减免额度 {discount_amount:.2f} 元；"
            f"用户可在商品列表中下单并消耗信用额度；备注：{loan.review_note}"
        )
        title = "后台审批通过"
        event_type = "ADMIN_APPROVED"
    else:
        loan.status = "REJECTED"
        loan.credit_limit = 0
        loan.approved_credit_limit = 0
        loan.approval_discount_amount = 0
        loan.order_discount_amount = 0
        loan.fee_rate = DEFAULT_FEE_RATE
        loan.fee_amount = 0
        loan.term_days = None
        loan.product_term_days = None
        loan.review_note = (req.review_note or "暂未通过当前授信审核").strip()
        loan.approved_at = None
        loan.due_date = None
        loan.disbursed_at = None
        loan.penalty_amount = 0
        loan.repaid_amount = 0
        loan.reduction_amount = 0
        loan.other_fee_amount = 0
        loan.paid_penalty_amount = 0
        loan.reduced_penalty_amount = 0
        loan.actual_repayment_date = None
        loan.repay_attempt_count = 0
        loan.collection_admin_id = None
        loan.collection_transferred_at = None
        loan.product_id = None
        loan.product_name = None
        loan.rights_title = None
        loan.rights_desc = None
        loan.rights_contact_phone = None
        loan.rights_price = 0
        loan.ecard_face_value = 0
        loan.product_total_price = 0
        loan.ecard_account = None
        loan.ecard_password = None
        loan.ecard_expires_at = None
        owner.approved_limit = 0
        _set_available_credit(owner, 0)

        detail = f"审批拒绝；备注：{loan.review_note}"
        title = "后台审批拒绝"
        event_type = "ADMIN_REJECTED"

    await log_user_event_async(
        db,
        user=owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type=event_type,
        title=title,
        detail=detail,
    )

    await db.commit()
    await db.refresh(loan)
    return serialize_loan(loan)

async def _update_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req: LoanUpdateRequest):
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner), joinedload(Loan.installments)).where(Loan.id == loan_id))
    ).unique().scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")

    payload = req.model_dump(exclude_unset=True)
    if not payload:
        return serialize_loan(loan)

    if set(payload.keys()) <= {"review_note"}:
        ensure_any_admin_page_permission(
            current_admin,
            ("users", "applications", "disbursements", "repayments", "collections", "financials"),
        )
    else:
        ensure_admin_page_permission(current_admin, "disbursements")

    owner = loan.owner
    change_messages = []
    previous_collection_note = loan.collection_note
    previous_review_note = loan.review_note
    fee_rate_updated = False

    if "status" in payload:
        new_status = payload["status"]
        if new_status not in LOAN_STATUSES:
            raise HTTPException(status_code=400, detail="订单状态非法")

        loan.status = new_status
        change_messages.append(f"状态调整为 {new_status}")

        if new_status == "APPROVED" and loan.approved_at is None:
            loan.approved_at = datetime.now()
        if new_status == "REJECTED":
            owner.approved_limit = 0
            _set_available_credit(owner, 0)
            loan.approved_at = None
            loan.approved_credit_limit = 0
            loan.credit_limit = 0 if "credit_limit" not in payload else loan.credit_limit
            loan.fee_rate = DEFAULT_FEE_RATE
            loan.fee_amount = 0
            loan.repay_attempt_count = 0
            loan.product_id = None
            loan.product_name = None
            loan.rights_title = None
            loan.rights_desc = None
            loan.rights_contact_phone = None
            loan.rights_price = 0
            loan.ecard_face_value = 0
            loan.product_total_price = 0
            loan.product_term_days = None
            loan.ecard_account = None
            loan.ecard_password = None
            loan.ecard_expires_at = None
        if new_status == "DISBURSED" and loan.disbursed_at is None:
            loan.disbursed_at = datetime.now()
        if new_status == "SETTLED":
            loan.penalty_amount = payload.get("penalty_amount", loan.penalty_amount)
            loan.repay_attempt_count = 0
        if new_status == "OVERDUE":
            await blacklist_user(
                db,
                owner,
                source="OVERDUE",
                reason="订单逾期自动进入黑名单",
                created_by=current_admin.username,
            )
            owner.approved_limit = 0
            _set_available_credit(owner, 0)
            owner.overdue_credit_locked = True
            loan.approved_credit_limit = 0

    if "credit_limit" in payload:
        loan.credit_limit = float(payload["credit_limit"])
        if loan.status == "APPROVED":
            loan.approved_credit_limit = float(payload["credit_limit"])
            _set_available_credit(owner, payload["credit_limit"])
            owner.overdue_credit_locked = False
        owner.approved_limit = int(payload["credit_limit"])
        change_messages.append(f"额度改为 {loan.credit_limit:.0f} 元")

    if "fee_rate" in payload:
        loan.fee_rate = normalize_fee_rate(payload["fee_rate"])
        fee_rate_updated = True
        change_messages.append(f"总费率改为 {loan.fee_rate * 100:.0f}%")

    if "term_days" in payload:
        loan.term_days = ensure_valid_term_days(payload["term_days"])
        change_messages.append(f"期限改为 {loan.term_days} 天")

    if "due_date" in payload and not loan.disbursed_at:
        loan.due_date = payload["due_date"]
        change_messages.append("重设还款日")

    if "penalty_amount" in payload:
        loan.penalty_amount = float(payload["penalty_amount"])
        change_messages.append(f"违约金改为 {loan.penalty_amount:.0f} 元")

    if "review_note" in payload:
        loan.review_note = payload["review_note"]
        change_messages.append("保存审批备注")

    if "collection_note" in payload:
        loan.collection_note = payload["collection_note"]
        change_messages.append("更新催收备注")

    if "credit_limit" in payload or fee_rate_updated:
        sync_loan_fee_fields(loan)

    if loan.disbursed_at and loan.term_days:
        loan.due_date = calculate_due_date(loan.disbursed_at, loan.term_days)
    elif loan.status == "DISBURSED" and loan.term_days:
        reference_time = datetime.now()
        loan.disbursed_at = reference_time
        loan.due_date = calculate_due_date(reference_time, loan.term_days)

    if (
        float(loan.repaid_amount or 0) + float(loan.reduction_amount or 0)
        > calculate_total_repayment_amount(loan) + 1e-6
    ):
        raise HTTPException(status_code=400, detail="已还款额与减免金额合计不能超过总还款额")

    if "collection_note" in payload:
        note_text = (payload.get("collection_note") or "").strip()
        previous_note = (previous_collection_note or "").strip()
        if note_text and note_text != previous_note:
            await log_user_event_async(
                db,
                user=owner,
                loan=loan,
                actor_type="ADMIN",
                operator_name=current_admin.username,
                event_type="ADMIN_COLLECTION_NOTE",
                title="新增催收备注",
                detail=note_text,
            )

    if "review_note" in payload:
        note_text = (payload.get("review_note") or "").strip()
        previous_note = (previous_review_note or "").strip()
        if note_text and note_text != previous_note:
            await log_user_event_async(
                db,
                user=owner,
                loan=loan,
                actor_type="ADMIN",
                operator_name=current_admin.username,
                event_type="ADMIN_REVIEW_NOTE",
                title="新增审批备注",
                detail=note_text,
            )

    await log_user_event_async(
        db,
        user=owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_LOAN_UPDATED",
        title="后台更新订单信息",
        detail="；".join(change_messages),
    )

    await db.commit()
    await db.refresh(loan)
    return serialize_loan(loan)

async def _disburse_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req: DisburseRequest):
    ensure_admin_page_permission(current_admin, "disbursements")
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner), joinedload(Loan.installments)).where(Loan.id == loan_id))
    ).unique().scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status != "WITHDRAWING":
        raise HTTPException(status_code=400, detail="仅待发卡订单支持发卡")
    await refresh_user_blacklist_status(db, loan.owner)
    if loan.owner.blacklist_hit:
        raise HTTPException(status_code=400, detail="该用户命中黑名单，当前不允许发卡")

    term_days = ensure_valid_term_days(req.term_days or loan.product_term_days or loan.term_days)
    if not term_days:
        raise HTTPException(status_code=400, detail="请先确认账期天数")
    if round_money(loan.product_total_price or 0) <= 0:
        raise HTTPException(status_code=400, detail="请先确认商品信息")

    now = datetime.now()
    is_cash_loan = (getattr(loan, "product_type", None) == "CASH_LOAN") or bool(getattr(loan, "total_repayment_amount_snapshot", 0))
    today = now.date()
    tomorrow_start = datetime(now.year, now.month, now.day) + timedelta(days=1)
    ecard_face_value = round_money(loan.ecard_face_value or 0)
    ecard_items = []
    if not is_cash_loan and ecard_face_value > 0:
        ecard_candidates = _extract_scalar_items(
            await db.execute(
                select(EcardPool)
                .where(
                    EcardPool.status == "AVAILABLE",
                    EcardPool.face_value <= ecard_face_value,
                    EcardPool.expires_at >= tomorrow_start,
                )
                .order_by(EcardPool.expires_at.asc(), EcardPool.id.asc())
            )
        )
        # 发卡允许多张E卡组合，但总面额必须与订单所需E卡金额完全一致。
        ecard_candidates = [item for item in ecard_candidates if item.expires_at.date() > today]
        ecard_items = _select_ecard_combo(ecard_candidates, ecard_face_value)
        if not ecard_items:
            raise HTTPException(status_code=400, detail=f"卡池库存不足：未找到面额 {ecard_face_value:.2f} 元且有效的京东E卡")

    disbursed_at = now
    loan.status = "DISBURSED"
    loan.term_days = term_days
    loan.product_term_days = term_days
    loan.credit_limit = round_money(loan.nominal_loan_amount or loan.product_total_price or ecard_face_value)
    loan.nominal_loan_amount = loan.credit_limit
    loan.upfront_fee_amount = round_money(loan.upfront_fee_amount or loan.fee_amount)
    loan.actual_disbursement_amount = round_money(loan.actual_disbursement_amount or max(loan.nominal_loan_amount - loan.upfront_fee_amount, 0))
    loan.total_repayment_amount_snapshot = round_money(loan.total_repayment_amount_snapshot or loan.nominal_loan_amount)
    if round_money(getattr(loan, "daily_overdue_fee_snapshot", 0)) <= 0 and getattr(loan, "product_id", None):
        loan_product = (await db.execute(select(Product).where(Product.id == loan.product_id))).scalar_one_or_none()
        if loan_product is not None:
            loan.daily_overdue_fee_snapshot = round_money(getattr(loan_product, "daily_overdue_fee", 0))
    if not loan.product_total_price:
        loan.product_total_price = round_money(loan.credit_limit + (loan.rights_price or loan.fee_amount))
    loan.disbursed_at = disbursed_at
    loan.interest_start_day = int(req.interest_start_day or getattr(loan, "interest_start_day", 1) or 1)
    loan.repayment_due_day = int(req.repayment_due_day or getattr(loan, "repayment_due_day", term_days) or term_days)
    loan.due_date = calculate_due_date(disbursed_at, loan.repayment_due_day)
    loan.penalty_amount = 0
    loan.repaid_amount = 0
    loan.reduction_amount = 0
    loan.other_fee_amount = 0
    loan.paid_penalty_amount = 0
    loan.reduced_penalty_amount = 0
    loan.actual_repayment_date = None
    loan.collection_admin_id = None
    loan.collection_transferred_at = None

    if is_cash_loan:
        momo_transaction = await create_or_get_momo_transaction_async(
            db,
            loan_id=loan.id,
            user_id=loan.user_id,
            transaction_type="DISBURSEMENT",
            phone=loan.owner.phone,
            amount=loan.actual_disbursement_amount,
            idempotency_key=f"DISBURSEMENT:{loan.id}",
        )
        if momo_transaction.status == "SUCCESS":
            loan.momo_disbursement_reference = momo_transaction.provider_reference
            transfer = None
        else:
            transfer = await momo_provider.disburse(loan.owner.phone, loan.actual_disbursement_amount, loan.id)
            complete_momo_transaction(
                momo_transaction,
                success=transfer.success,
                reference=transfer.reference,
                provider=transfer.provider,
                message=transfer.message,
            )
            if transfer.success:
                loan.momo_disbursement_reference = transfer.reference
        if momo_transaction.status != "SUCCESS" and not transfer.success:
            await db.commit()
            raise HTTPException(status_code=400, detail=transfer.message or "MoMo放款失败")
        loan.ecard_account = None
        loan.ecard_password = None
        loan.ecard_expires_at = None
    elif ecard_items:
        first_ecard = ecard_items[0]
        loan.ecard_account = first_ecard.account
        loan.ecard_password = first_ecard.password
        loan.ecard_expires_at = first_ecard.expires_at
        for item in ecard_items:
            item.status = "ASSIGNED"
            item.loan_id = loan.id
            item.assigned_at = now
            db.add(
                LoanEcard(
                    loan_id=loan.id,
                    ecard_pool_id=item.id,
                    account=item.account,
                    password=item.password,
                    face_value=round_money(item.face_value),
                    expires_at=item.expires_at,
                )
            )
    else:
        loan.ecard_account = None
        loan.ecard_password = None
        loan.ecard_expires_at = None

    await ensure_installment_records_async(db, loan)
    await create_disbursement_transaction_async(
        db,
        loan,
        operator_name=current_admin.username,
        note="后台确认MoMo放款" if is_cash_loan else "后台确认发放京东E卡",
    )

    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_CASH_DISBURSED" if is_cash_loan else "ADMIN_CARD_ISSUED",
        title="后台确认MoMo放款" if is_cash_loan else "后台确认发卡",
        detail=(
            f"贷款产品：{loan.product_name or '未命名产品'}；"
            f"名义借款 {round_money(loan.nominal_loan_amount):.2f}；"
            f"上扣费用 {round_money(loan.upfront_fee_amount):.2f}；"
            f"实际到账 {round_money(loan.actual_disbursement_amount):.2f}；"
            f"MoMo流水 {loan.momo_disbursement_reference or '-'}；"
            f"账期 {loan.term_days} 天；"
            f"到期日 {loan.due_date.strftime('%Y-%m-%d')}；"
            f"卡池记录 {','.join(f'#{item.id}' for item in ecard_items) if ecard_items else '无E卡'}。"
        ),
    )

    await db.commit()
    await db.refresh(loan)
    return {"msg": "发卡成功", "loan": serialize_loan(loan)}

async def _reject_card_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req: LoanFollowUpRequest):
    ensure_admin_page_permission(current_admin, "disbursements")
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status != "WITHDRAWING":
        raise HTTPException(status_code=400, detail="仅待发卡订单支持拒绝发卡")
    loan.status = "CARD_REJECTED"
    loan.review_note = (req.note or "后台拒绝发卡").strip()
    loan.disbursed_at = None
    loan.due_date = None
    loan.ecard_account = None
    loan.ecard_password = None
    loan.ecard_expires_at = None
    if loan.is_extension_fee_order:
        _set_available_credit(
            loan.owner,
            round_money(getattr(loan.owner, "available_credit_limit", 0)) + round_money(loan.product_total_price or loan.credit_limit or 0),
        )
    else:
        loan.owner.approved_limit = 0
        _set_available_credit(loan.owner, 0)
    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_CARD_REJECTED",
        title="后台拒绝发卡",
        detail=loan.review_note,
    )
    await db.commit()
    await db.refresh(loan)
    return {"msg": "已拒绝发卡", "loan": serialize_loan(loan)}


async def _reissue_card_loan(db: AsyncSession, current_admin: Admin, loan_id: int):
    ensure_admin_page_permission(current_admin, "users")
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status != "CARD_REJECTED":
        raise HTTPException(status_code=400, detail="仅拒发卡订单支持二次发卡")
    if loan.card_reissue_closed:
        raise HTTPException(status_code=400, detail="该订单已关闭二次发卡")
    await refresh_user_blacklist_status(db, loan.owner)
    if loan.owner.blacklist_hit:
        raise HTTPException(status_code=400, detail="该用户命中黑名单，不能二次发卡")
    loan.status = "WITHDRAWING"
    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_CARD_REISSUE",
        title="后台开启二次发卡",
        detail="该订单已重新进入待发卡列表。",
    )
    await db.commit()
    await db.refresh(loan)
    return {"msg": "已进入待发卡", "loan": serialize_loan(loan)}


def _return_loan_to_approved_for_reorder(loan: Loan) -> float:
    """将待发卡/拒发卡订单退回待下单。

    :param loan: 订单对象
    :return: 恢复后的可用额度
    """
    approved_credit_limit = round_money(loan.approved_credit_limit or loan.credit_limit or 0)
    owner = loan.owner

    loan.status = "APPROVED"
    loan.card_reissue_closed = False
    loan.credit_limit = approved_credit_limit
    loan.approved_credit_limit = approved_credit_limit
    loan.fee_rate = DEFAULT_FEE_RATE
    loan.fee_amount = 0
    loan.order_discount_amount = 0
    loan.due_date = None
    loan.disbursed_at = None
    loan.penalty_amount = 0
    loan.repaid_amount = 0
    loan.reduction_amount = 0
    loan.other_fee_amount = 0
    loan.paid_penalty_amount = 0
    loan.reduced_penalty_amount = 0
    loan.actual_repayment_date = None
    loan.reminder_count = 0
    loan.last_reminded_at = None
    loan.collection_count = 0
    loan.last_collection_at = None
    loan.collection_note = None
    loan.collection_admin_id = None
    loan.collection_transferred_at = None
    loan.repay_attempt_count = 0

    # 退回待下单的本质是撤销本次错误商品选择，让用户重新选购。
    loan.product_id = None
    loan.product_name = None
    loan.rights_title = None
    loan.rights_desc = None
    loan.rights_contact_phone = None
    loan.rights_price = 0
    loan.ecard_face_value = 0
    loan.product_total_price = 0
    loan.product_term_days = loan.term_days
    loan.ecard_account = None
    loan.ecard_password = None
    loan.ecard_expires_at = None
    loan.order_no = ""

    if owner:
        owner.approved_limit = int(approved_credit_limit)
        _set_available_credit(owner, approved_credit_limit)
        owner.overdue_credit_locked = False
    return approved_credit_limit


async def _close_card_reissue(db: AsyncSession, current_admin: Admin, loan_id: int):
    ensure_any_admin_page_permission(current_admin, ("users", "disbursements"))
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status not in {"WITHDRAWING", "CARD_REJECTED"}:
        raise HTTPException(status_code=400, detail="仅待发卡/拒发卡订单支持退回待下单")
    if loan.is_extension_fee_order:
        raise HTTPException(status_code=400, detail="展期权益订单不支持退回待下单")
    approved_credit_limit = _return_loan_to_approved_for_reorder(loan)
    loan.review_note = "退回待下单：原下单信息有误，用户可重新选择商品下单"
    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_CARD_RETURN_TO_ORDER",
        title="后台退回待下单",
        detail=f"已撤销原待发卡商品信息，恢复可用额度 {approved_credit_limit:.2f} 元，用户可重新下单。",
    )
    await db.commit()
    await db.refresh(loan)
    return {"msg": "已退回待下单", "loan": serialize_loan(loan)}


async def _extend_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req):
    ensure_any_admin_page_permission(current_admin, ("repayments", "collections"))
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner), joinedload(Loan.installments)).where(Loan.id == loan_id))
    ).unique().scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status not in {"DISBURSED", "OVERDUE"}:
        raise HTTPException(status_code=400, detail="当前账单不支持展期")
    days = int(req.days)
    fee_order = None
    if req.extension_type == "FREE":
        reduction_amount = round_money(req.reduction_amount or 0)
        if reduction_amount > 0:
            await ensure_installment_records_async(db, loan)
            await register_reduction_async(
                db,
                loan,
                reduction_amount,
                operator_name=current_admin.username,
                note=(req.note or "无附加条件展期减免平账").strip(),
            )
    else:
        ready_orders = await _find_ready_fee_extension_orders(db, loan)
        if req.fee_order_id:
            fee_order = next((item for item in ready_orders if int(item.id) == int(req.fee_order_id)), None)
        else:
            fee_order = ready_orders[0] if ready_orders else None
        if not fee_order:
            raise HTTPException(status_code=400, detail="带息费展期前，用户必须先完成一笔未使用的纯权益订单")
    base_due_date = loan.due_date or datetime.now()
    old_due_date = loan.due_date
    loan.due_date = base_due_date + timedelta(days=days)
    if old_due_date and loan.installments:
        for installment in loan.installments:
            if getattr(installment, "status", None) != "SETTLED":
                installment.due_date = installment.due_date + timedelta(days=days)
    loan.status = "DISBURSED"
    loan.overdue_hidden = True
    loan.extension_count = int(loan.extension_count or 0) + 1
    loan.extension_type = req.extension_type
    loan.extension_note = (req.note or "").strip() or None
    if fee_order:
        fee_order.due_date = loan.due_date
        fee_order.extension_used_at = datetime.now()
        fee_order.extension_type = "FEE"
        fee_order.extension_note = f"作为订单 {loan.id} 带息费展期权益订单"
        await ensure_installment_records_async(db, fee_order)
        if fee_order.installments:
            for installment in fee_order.installments:
                if getattr(installment, "status", None) != "SETTLED":
                    installment.due_date = loan.due_date
    sync_loan_repayment_state(loan)
    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_LOAN_EXTENDED",
        title="后台账单展期",
        detail=f"{'无附加条件' if req.extension_type == 'FREE' else '带息费'}展期 {days} 天；备注：{loan.extension_note or '--'}",
    )
    await db.commit()
    await db.refresh(loan)
    return {"msg": "展期成功", "loan": serialize_loan(loan)}


async def _adjust_available_credit(db: AsyncSession, current_admin: Admin, loan_id: int, req: AvailableCreditAdjustRequest):
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status == "OVERDUE" and not loan.overdue_hidden:
        raise HTTPException(status_code=400, detail="逾期用户需先关闭逾期状态显示，再增加可用信用额度")
    if loan.status == "OVERDUE" or loan.overdue_hidden:
        ensure_admin_page_permission(current_admin, "collections")
    else:
        ensure_admin_page_permission(current_admin, "applications")
    owner = loan.owner
    before_amount = round_money(getattr(owner, "available_credit_limit", 0))
    add_amount = round_money(req.amount)
    _set_available_credit(owner, before_amount + add_amount)
    if loan.status != "OVERDUE" or loan.overdue_hidden:
        owner.overdue_credit_locked = False
    await log_user_event_async(
        db,
        user=owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_AVAILABLE_CREDIT_ADJUSTED",
        title="后台增加可用信用额度",
        detail=f"增加 {add_amount:.2f} 元；调整前 {before_amount:.2f} 元；调整后 {owner.available_credit_limit:.2f} 元；备注：{(req.note or '--').strip()}",
    )
    await db.commit()
    await db.refresh(loan)
    return {"msg": "可用额度已增加", "loan": serialize_loan(loan)}


async def _set_approved_credit_limit(db: AsyncSession, current_admin: Admin, loan_id: int, req: ApprovedCreditSetRequest):
    """调低已审批但未下单用户的额度。

    :param db: 异步数据库会话
    :param current_admin: 当前后台用户
    :param loan_id: 订单ID
    :param req: 调整后额度请求体
    :return: 调整结果
    """
    ensure_admin_page_permission(current_admin, "applications")
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status != "APPROVED":
        raise HTTPException(status_code=400, detail="仅支持调整已审批且尚未下单的额度")

    owner = loan.owner
    before_limit = round_money(getattr(loan, "approved_credit_limit", 0) or getattr(loan, "credit_limit", 0) or 0)
    next_limit = round_money(req.credit_limit)
    if next_limit > before_limit + 1e-6:
        raise HTTPException(status_code=400, detail="此操作仅支持调减额度，增加额度请使用增加可用额度")

    # 未下单阶段没有账单金额，调减时需同步审批额度与用户可用额度，避免前后台显示不一致。
    loan.approved_credit_limit = next_limit
    loan.credit_limit = next_limit
    owner.approved_limit = int(next_limit)
    _set_available_credit(owner, next_limit)
    owner.overdue_credit_locked = False

    await log_user_event_async(
        db,
        user=owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_APPROVED_CREDIT_SET",
        title="后台调减授信额度",
        detail=f"调整前 {before_limit:.2f} 元；调整后 {next_limit:.2f} 元；备注：{(req.note or '--').strip()}",
    )
    await db.commit()
    await db.refresh(loan)
    return {"msg": "审批额度已调整", "loan": serialize_loan(loan)}


async def _update_overdue_display(db: AsyncSession, current_admin: Admin, loan_id: int, req: OverdueDisplayRequest):
    ensure_admin_page_permission(current_admin, "collections")
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner), joinedload(Loan.installments)).where(Loan.id == loan_id))
    ).unique().scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status not in {"OVERDUE", "DISBURSED"}:
        raise HTTPException(status_code=400, detail="当前账单不支持调整逾期显示")
    loan.overdue_hidden = bool(req.overdue_hidden)
    if loan.overdue_hidden and loan.status == "OVERDUE":
        loan.status = "DISBURSED"
    if loan.overdue_hidden:
        loan.owner.overdue_credit_locked = False
    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_OVERDUE_DISPLAY_UPDATED",
        title="后台调整逾期状态显示",
        detail=f"逾期显示已{'关闭' if loan.overdue_hidden else '开启'}；备注：{(req.note or '--').strip()}",
    )
    await db.commit()
    await db.refresh(loan)
    return {"msg": "逾期状态已调整", "loan": serialize_loan(loan)}


async def _get_blacklist_entries(db: AsyncSession, current_admin: Admin, keyword: Optional[str], skip: int, limit: int):
    ensure_admin_page_permission(current_admin, "blacklist")
    limit = min(max(limit, 1), 100)
    stmt = select(BlacklistEntry).where(BlacklistEntry.removed_at.is_(None))
    if keyword:
        pattern = f"%{keyword.strip()}%"
        stmt = stmt.where(
            or_(
                BlacklistEntry.name.like(pattern),
                BlacklistEntry.phone.like(pattern),
                BlacklistEntry.id_card_num.like(pattern),
            )
        )
    total = (await db.scalar(select(func.count()).select_from(stmt.subquery()))) or 0
    items = (
        await db.execute(stmt.order_by(BlacklistEntry.created_at.desc()).offset(skip).limit(limit))
    ).scalars().all()
    return {"total": total, "page": skip // limit + 1, "size": limit, "items": [serialize_blacklist_entry(item) for item in items]}


def serialize_overdue_fee_config(item: OverdueFeeConfig):
    return {
        "id": item.id,
        "daily_penalty_amount": round_money(item.daily_penalty_amount),
        "effective_date": item.effective_date,
        "note": item.note,
        "created_by": item.created_by,
        "created_at": item.created_at,
    }


async def _get_overdue_fee_configs(db: AsyncSession, current_admin: Admin, skip: int, limit: int):
    ensure_admin_page_permission(current_admin, "overdue-config")
    limit = min(max(limit, 1), 100)
    total = (await db.scalar(select(func.count(OverdueFeeConfig.id)))) or 0
    items = (
        await db.execute(
            select(OverdueFeeConfig)
            .order_by(OverdueFeeConfig.effective_date.desc(), OverdueFeeConfig.id.desc())
            .offset(skip)
            .limit(limit)
        )
    ).scalars().all()
    return {"total": total, "page": skip // limit + 1, "size": limit, "items": [serialize_overdue_fee_config(item) for item in items]}


async def _create_overdue_fee_config(db: AsyncSession, current_admin: Admin, req: OverdueFeeConfigCreateRequest):
    ensure_admin_page_permission(current_admin, "overdue-config")
    existed = (
        await db.execute(select(OverdueFeeConfig).where(OverdueFeeConfig.effective_date == req.effective_date))
    ).scalars().first()
    if existed:
        raise HTTPException(status_code=400, detail="该生效日已存在逾期费用配置")
    item = OverdueFeeConfig(
        daily_penalty_amount=round_money(req.daily_penalty_amount),
        effective_date=req.effective_date,
        note=(req.note or "").strip() or None,
        created_by=current_admin.username,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return serialize_overdue_fee_config(item)


async def _manual_blacklist_user(db: AsyncSession, current_admin: Admin, user_id: int, req: LoanFollowUpRequest):
    ensure_any_admin_page_permission(current_admin, ("users", "applications", "disbursements", "repayments", "collections", "financials"))
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    entry = await blacklist_user(
        db,
        user,
        source="MANUAL",
        reason=(req.note or "后台一键拉黑").strip(),
        created_by=current_admin.username,
    )
    await log_user_event_async(
        db,
        user=user,
        loan=await get_latest_loan_async(db, user.id),
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_BLACKLIST_USER",
        title="后台一键拉黑",
        detail=entry.reason or "后台一键拉黑",
    )
    await db.commit()
    return {"msg": "已加入黑名单", "entry": serialize_blacklist_entry(entry)}


async def _remove_blacklist_user(db: AsyncSession, current_admin: Admin, user_id: int, req: LoanFollowUpRequest):
    ensure_admin_page_permission(current_admin, "blacklist")
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    count = await remove_user_from_blacklist(db, user, removed_by=current_admin.username, reason=req.note)
    await log_user_event_async(
        db,
        user=user,
        loan=await get_latest_loan_async(db, user.id),
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_BLACKLIST_REMOVE",
        title="后台移出黑名单",
        detail=req.note or f"已移出 {count} 条命中记录。",
    )
    await db.commit()
    return {"msg": "已移出黑名单", "removed": count}


async def _upload_blacklist(db: AsyncSession, current_admin: Admin, file: UploadFile):
    ensure_admin_page_permission(current_admin, "blacklist")
    result = await upload_blacklist_entries(db, file, created_by=current_admin.username)
    await db.commit()
    return result

async def _settle_loan(db: AsyncSession, current_admin: Admin, loan_id: int):
    ensure_admin_page_permission(current_admin, "financials")
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status not in {"DISBURSED", "OVERDUE"}:
        raise HTTPException(status_code=400, detail="订单状态不支持结清")

    remaining_amount = calculate_remaining_repayment_amount(loan)
    if remaining_amount > 0:
        await register_repayment_async(
            db,
            loan,
            remaining_amount,
            operator_name=current_admin.username,
            note="后台一键结清补录剩余待还金额",
            transaction_type="SETTLEMENT",
        )
    sync_loan_repayment_state(loan)
    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_SETTLED",
        title="后台确认结清",
        detail="后台已登记该订单完成还款结清。",
    )

    await db.commit()
    return {"msg": "结清成功"}

async def _finance_reconcile_loan(
    db: AsyncSession,
    current_admin: Admin,
    loan_id: int,
    req: LoanFinanceReconcileRequest,
):
    ensure_admin_page_permission(current_admin, "financials")
    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status not in {"DISBURSED", "OVERDUE"}:
        raise HTTPException(status_code=400, detail="当前订单暂不支持财务平账")
    was_overdue_or_locked = loan.status == "OVERDUE" or bool(getattr(loan.owner, "overdue_credit_locked", False))

    received_amount = round(float(req.received_amount or 0), 2)
    reduction_amount = round(float(req.reduction_amount or 0), 2)
    extra_fee_amount = round(float(req.other_fee_amount or 0), 2)
    actual_repayment_date = _normalize_actual_repayment_date(req.actual_repayment_date)
    note = (req.note or "").strip()
    if received_amount <= 0 and reduction_amount <= 0 and extra_fee_amount <= 0:
        raise HTTPException(status_code=400, detail="请填写登记收款、减免金额或其他费用")

    # 逾期费按“实际还款日”冻结，避免登记日晚于客户付款日时继续累加。
    daily_overdue_fee_snapshot = round_money(getattr(loan, "daily_overdue_fee_snapshot", 0))
    if daily_overdue_fee_snapshot > 0:
        overdue_days = calculate_overdue_days(loan.due_date, actual_repayment_date)
        penalty_meta = {
            "overdue_days": overdue_days,
            "daily_penalty_amount": daily_overdue_fee_snapshot,
            "penalty_amount": round(overdue_days * daily_overdue_fee_snapshot, 2),
        }
    else:
        # 历史订单没有快照时，兼容使用按日期生效的全局逾期标准。
        penalty_meta = await calculate_penalty_by_repayment_date(db, loan.due_date, actual_repayment_date)
    loan.actual_repayment_date = actual_repayment_date
    loan.penalty_amount = penalty_meta["penalty_amount"]

    unpaid_penalty_amount = round(
        max(
            float(loan.penalty_amount or 0)
            - float(loan.paid_penalty_amount or 0)
            - float(loan.reduced_penalty_amount or 0),
            0,
        ),
        2,
    )
    extra_fee_split = split_extra_fee_for_penalty(extra_fee_amount, unpaid_penalty_amount)
    penalty_paid_now = extra_fee_split["penalty_paid_now"]
    other_fee_amount = extra_fee_split["other_fee_amount"]
    effective_repayment_amount = round(received_amount + penalty_paid_now, 2)

    total_amount = calculate_total_repayment_amount(loan)
    next_repaid_amount = round(float(loan.repaid_amount or 0) + effective_repayment_amount, 2)
    next_reduction_amount = round(float(loan.reduction_amount or 0) + reduction_amount, 2)
    if next_repaid_amount + next_reduction_amount > total_amount + 1e-6:
        raise HTTPException(status_code=400, detail="收款金额与减免金额累计不能超过总还款额")

    await ensure_installment_records_async(db, loan)

    if effective_repayment_amount > 0:
        await register_repayment_async(
            db,
            loan,
            effective_repayment_amount,
            operator_name=current_admin.username,
            note=note or "后台登记收款",
        )

    if reduction_amount > 0:
        await register_reduction_async(
            db,
            loan,
            reduction_amount,
            operator_name=current_admin.username,
            note=note or "后台登记减免",
        )
    if other_fee_amount > 0:
        await register_other_fee_async(
            db,
            loan,
            other_fee_amount,
            operator_name=current_admin.username,
            note=note or "后台登记其他费用",
        )

    remaining_amount = calculate_remaining_repayment_amount(loan)
    sync_loan_repayment_state(loan)
    if received_amount > 0 and not was_overdue_or_locked:
        before_available = round_money(getattr(loan.owner, "available_credit_limit", 0))
        _set_available_credit(loan.owner, before_available + received_amount)

    detail_parts = []
    if received_amount > 0:
        detail_parts.append(f"登记收款 {received_amount:.2f} 元")
    if penalty_paid_now > 0:
        detail_parts.append(f"额外收款冲抵逾期费 {penalty_paid_now:.2f} 元")
    if reduction_amount > 0:
        detail_parts.append(f"登记减免 {reduction_amount:.2f} 元")
    if other_fee_amount > 0:
        detail_parts.append(f"登记其他费用 {other_fee_amount:.2f} 元")
    detail_parts.append(f"实际还款日 {actual_repayment_date.isoformat()}")
    detail_parts.append(f"逾期 {penalty_meta['overdue_days']} 天")
    detail_parts.append(f"日逾期费标准 {penalty_meta['daily_penalty_amount']:.2f} 元")
    detail_parts.append(f"应收逾期费 {float(loan.penalty_amount or 0):.2f} 元")
    detail_parts.append(f"累计已还 {loan.repaid_amount:.2f} 元")
    detail_parts.append(f"累计减免 {loan.reduction_amount:.2f} 元")
    detail_parts.append(f"累计已收逾期费 {float(loan.paid_penalty_amount or 0):.2f} 元")
    detail_parts.append(f"累计其他费用 {float(loan.other_fee_amount or 0):.2f} 元")
    detail_parts.append(f"剩余待还 {remaining_amount:.2f} 元")
    if note:
        detail_parts.append(f"备注：{note}")
    if loan.status == "SETTLED":
        detail_parts.append("订单已完成平账结清")

    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_FINANCE_RECONCILE",
        title="财务登记平账",
        detail="；".join(detail_parts),
    )

    await db.commit()
    await db.refresh(loan)
    return serialize_loan(loan)

async def _remind_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req: LoanFollowUpRequest):
    ensure_any_admin_page_permission(current_admin, ("repayments", "collections", "message-center"))
    loan = (
        await db.execute(
            select(Loan)
            .options(joinedload(Loan.owner), joinedload(Loan.review_admin))
            .where(Loan.id == loan_id)
        )
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status not in {"DISBURSED", "OVERDUE"}:
        raise HTTPException(status_code=400, detail="当前订单不需要登记提醒")
    if is_collection_stage(loan):
        raise HTTPException(status_code=400, detail=f"该订单逾期已超过 {COLLECTION_TRANSFER_OVERDUE_DAYS} 天，请在催收管理处理")
    if not is_super_admin(current_admin) and not admin_has_permission(current_admin, "message-center"):
        if int(loan.review_admin_id or 0) != int(current_admin.id):
            raise HTTPException(status_code=403, detail="仅可跟进分配给你的还款订单")

    loan.reminder_count = (loan.reminder_count or 0) + 1
    loan.last_reminded_at = datetime.now()
    note = (req.note or "已执行当日还款提醒").strip()

    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_REMIND",
        title="登记还款提醒",
        detail=f"第 {loan.reminder_count} 次提醒；备注：{note}",
    )

    await db.commit()
    await db.refresh(loan)
    return serialize_loan(loan)

async def _collect_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req: LoanFollowUpRequest):
    ensure_admin_page_permission(current_admin, "collections")
    loan = (
        await db.execute(
            select(Loan)
            .options(joinedload(Loan.owner), joinedload(Loan.collection_admin))
            .where(Loan.id == loan_id)
        )
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    if loan.status != "OVERDUE":
        raise HTTPException(status_code=400, detail="仅逾期订单支持登记催收")
    if not is_collection_stage(loan):
        raise HTTPException(status_code=400, detail=f"逾期超过 {COLLECTION_TRANSFER_OVERDUE_DAYS} 天后才可进入催收")

    await assign_collection_admin_if_needed_async(db, loan)
    if not is_super_admin(current_admin):
        if int(loan.collection_admin_id or 0) != int(current_admin.id):
            raise HTTPException(status_code=403, detail="仅可跟进分配给你的催收订单")

    loan.collection_count = (loan.collection_count or 0) + 1
    loan.last_collection_at = datetime.now()
    loan.collection_note = (req.note or "已执行逾期催收").strip()

    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_COLLECT",
        title="登记催收跟进",
        detail=f"第 {loan.collection_count} 次催收；备注：{loan.collection_note}",
    )

    await db.commit()
    await db.refresh(loan)
    return serialize_loan(loan)

async def _ack_repay_attempt(db: AsyncSession, current_admin: Admin, loan_id: int):
    ensure_any_admin_page_permission(current_admin, ("repayments", "collections"))
    loan = (
        await db.execute(
            select(Loan)
            .options(
                joinedload(Loan.owner),
                joinedload(Loan.review_admin),
                joinedload(Loan.collection_admin),
            )
            .where(Loan.id == loan_id)
        )
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")
    ensure_stage_access_for_admin(current_admin, loan)

    cleared_count = int(loan.repay_attempt_count or 0)
    if cleared_count > 0:
        loan.repay_attempt_count = 0
        await log_user_event_async(
            db,
            user=loan.owner,
            loan=loan,
            actor_type="ADMIN",
            operator_name=current_admin.username,
            event_type="ADMIN_REPAY_ATTEMPT_ACK",
            title="查看还款跟进",
            detail=f"后台查看跟进时已清除还款点击提醒 {cleared_count} 次。",
        )
        await db.commit()
        await db.refresh(loan)

    return {
        "loan_id": loan.id,
        "cleared_count": cleared_count,
        "repay_attempt_count": int(loan.repay_attempt_count or 0),
    }

async def _get_loan_assignees(db: AsyncSession, current_admin: Admin, stage: str):
    if not is_super_admin(current_admin):
        raise HTTPException(status_code=403, detail="仅超级管理员可查看可分配人员")

    normalized_stage = (stage or "").strip().lower()
    if normalized_stage not in {"review", "collection"}:
        raise HTTPException(status_code=400, detail="分配阶段参数非法")

    role_key = "REVIEW" if normalized_stage == "review" else "COLLECTION"
    assignees = await list_admins_by_role_async(db, role_key)
    return [{"id": item.id, "username": item.username} for item in assignees]

async def _assign_loan(db: AsyncSession, current_admin: Admin, loan_id: int, req: LoanAssignRequest):
    stage = (req.stage or "").strip().lower()
    if stage not in {"review", "collection"}:
        raise HTTPException(status_code=400, detail="分配阶段参数非法")

    roles = current_admin_roles(current_admin)
    is_admin = "ADMIN" in roles
    is_review_takeover = (
        not is_admin
        and stage == "review"
        and "REVIEW" in roles
        and int(req.admin_id) == int(current_admin.id)
        and admin_has_permission(current_admin, "loan-review-takeover")
    )
    if not is_admin and not is_review_takeover:
        raise HTTPException(status_code=403, detail="仅超级管理员可手动改派订单")

    loan = (
        await db.execute(select(Loan).options(joinedload(Loan.owner)).where(Loan.id == loan_id))
    ).scalar_one_or_none()
    if not loan:
        raise HTTPException(status_code=404, detail="订单不存在")

    if is_review_takeover and loan.status != "REVIEWING":
        raise HTTPException(status_code=400, detail="仅审核中的申请可转入自己")

    assignee = (await db.execute(select(Admin).where(Admin.id == req.admin_id))).scalar_one_or_none()
    if not assignee:
        raise HTTPException(status_code=404, detail="分配目标不存在")

    role_key = "REVIEW" if stage == "review" else "COLLECTION"
    if not admin_has_role(assignee, role_key):
        raise HTTPException(status_code=400, detail=f"目标账号不是{('审核员' if role_key == 'REVIEW' else '催收员')}")

    if stage == "review":
        previous = loan.review_admin_id
        loan.review_admin_id = assignee.id
        title = "审核员转入自己" if is_review_takeover else "超管手动改派审核负责人"
        detail = f"审核负责人由 #{previous or '-'} 调整为 #{assignee.id}（{assignee.username}）"
    else:
        if not is_admin:
            raise HTTPException(status_code=403, detail="仅超级管理员可手动改派催收订单")
        if not is_collection_stage(loan):
            raise HTTPException(status_code=400, detail=f"逾期超过 {COLLECTION_TRANSFER_OVERDUE_DAYS} 天后才可改派催收负责人")
        previous = loan.collection_admin_id
        loan.collection_admin_id = assignee.id
        if loan.collection_transferred_at is None:
            loan.collection_transferred_at = datetime.now()
        title = "超管手动改派催收负责人"
        detail = f"催收负责人由 #{previous or '-'} 调整为 #{assignee.id}（{assignee.username}）"

    await log_user_event_async(
        db,
        user=loan.owner,
        loan=loan,
        actor_type="ADMIN",
        operator_name=current_admin.username,
        event_type="ADMIN_ASSIGNMENT_UPDATED",
        title=title,
        detail=detail,
    )

    await db.commit()
    return {
        "loan_id": loan.id,
        "stage": stage,
        "assignee_id": assignee.id,
        "assignee_name": assignee.username,
    }

async def _get_admin_users(db: AsyncSession, current_admin: Admin, keyword: Optional[str], skip: int, limit: int):
    ensure_admin_page_permission(current_admin, "admin-users")
    limit = min(max(limit, 1), 100)

    stmt = select(Admin)
    if keyword:
        pattern = f"%{keyword.strip()}%"
        stmt = stmt.where(Admin.username.like(pattern))

    total = (await db.scalar(select(func.count()).select_from(stmt.subquery()))) or 0
    admins = (
        await db.execute(stmt.order_by(Admin.created_at.desc(), Admin.id.desc()).offset(skip).limit(limit))
    ).scalars().all()
    return {
        "total": total,
        "page": skip // limit + 1,
        "size": limit,
        "items": [serialize_admin_user(item, current_admin) for item in admins],
    }


async def _get_admin_audit_logs(
    db: AsyncSession,
    current_admin: Admin,
    keyword: Optional[str],
    actor_type: Optional[str],
    skip: int,
    limit: int,
    event_type: Optional[str] = None,
    object_type: Optional[str] = None,
    start_date=None,
    end_date=None,
):
    """查询管理员操作审计日志。

    :param db: 异步数据库会话
    :param current_admin: 当前管理员
    :param keyword: 搜索关键字
    :param actor_type: 操作者类型筛选
    :param skip: 分页偏移量
    :param limit: 每页数量
    :return: 审计日志分页数据
    """
    ensure_admin_page_permission(current_admin, "audit-log")
    limit = min(max(limit, 1), 100)
    stmt = (
        select(UserEvent, User.name, User.phone, Loan.order_no)
        .join(User, User.id == UserEvent.user_id)
        .outerjoin(Loan, Loan.id == UserEvent.loan_id)
    )
    if actor_type and actor_type.strip() and actor_type.strip().upper() != "ALL":
        stmt = stmt.where(UserEvent.actor_type == actor_type.strip().upper())
    if event_type and event_type.strip() and event_type.strip().upper() != "ALL":
        stmt = stmt.where(UserEvent.event_type == event_type.strip().upper())
    if object_type and object_type.strip().upper() == "USER":
        stmt = stmt.where(UserEvent.user_id.is_not(None))
    elif object_type and object_type.strip().upper() == "LOAN":
        stmt = stmt.where(UserEvent.loan_id.is_not(None))
    if start_date is not None:
        stmt = stmt.where(UserEvent.created_at >= start_date)
    if end_date is not None:
        stmt = stmt.where(UserEvent.created_at < end_date + timedelta(days=1))
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        stmt = stmt.where(
            or_(
                UserEvent.title.like(pattern),
                UserEvent.detail.like(pattern),
                UserEvent.operator_name.like(pattern),
                User.name.like(pattern),
                User.phone.like(pattern),
                Loan.order_no.like(pattern),
            )
        )
    total = (await db.scalar(select(func.count()).select_from(stmt.subquery()))) or 0
    rows = (
        await db.execute(stmt.order_by(UserEvent.created_at.desc(), UserEvent.id.desc()).offset(skip).limit(limit))
    ).all()
    return {
        "total": total,
        "page": skip // limit + 1,
        "size": limit,
        "items": [
            {
                "id": event.id,
                "user_id": event.user_id,
                "user_name": user_name,
                "user_phone": user_phone,
                "loan_id": event.loan_id,
                "loan_order_no": loan_order_no,
                "actor_type": event.actor_type,
                "operator_name": event.operator_name,
                "event_type": event.event_type,
                "title": event.title,
                "detail": event.detail,
                "created_at": event.created_at,
            }
            for event, user_name, user_phone, loan_order_no in rows
        ],
    }


def _resolve_kyc_review_flags(user: User) -> list[str]:
    flags = []
    real_name_status = (user.real_name_status or "").upper()
    face_auth_status = (user.face_auth_status or "").upper()
    if real_name_status not in {"VERIFIED", "AUTHED", "PASS", "PASSED"}:
        flags.append("实名未完成")
    if face_auth_status not in {"APPROVED", "PASS", "PASSED"}:
        flags.append("人脸未完成")
    if not user.id_card_num:
        flags.append("缺身份证号")
    if user.location_risk_blocked:
        flags.append("位置风控")
    if user.blacklist_hit:
        flags.append("黑名单")
    if user.risk_list_hit:
        flags.append("风险名单")
    return flags


async def _get_kyc_review_queue(db: AsyncSession, current_admin: Admin, keyword: Optional[str], skip: int, limit: int):
    """查询待人工复核的 KYC 用户队列。

    :param db: 异步数据库会话
    :param current_admin: 当前管理员
    :param keyword: 搜索关键字
    :param skip: 分页偏移量
    :param limit: 每页数量
    :return: KYC 复核分页数据
    """
    ensure_any_admin_page_permission(current_admin, ("users", "applications", "admin-users", "kyc-review"))
    limit = min(max(limit, 1), 100)
    stmt = (
        select(User, Channel.channel_name, Channel.sales_name)
        .outerjoin(Channel, Channel.id == User.source_channel_id)
        .where(
            or_(
                User.real_name_status.is_(None),
                ~func.upper(User.real_name_status).in_(["VERIFIED", "AUTHED", "PASS", "PASSED"]),
                User.face_auth_status.is_(None),
                ~func.upper(User.face_auth_status).in_(["APPROVED", "PASS", "PASSED"]),
                User.location_risk_blocked.is_(True),
                User.blacklist_hit.is_(True),
                User.risk_list_hit.is_(True),
            )
        )
    )
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        stmt = stmt.where(or_(User.phone.like(pattern), User.name.like(pattern), User.id_card_num.like(pattern)))

    total = (await db.scalar(select(func.count()).select_from(stmt.subquery()))) or 0
    rows = (
        await db.execute(stmt.order_by(User.created_at.desc(), User.id.desc()).offset(skip).limit(limit))
    ).all()
    items = []
    for user, channel_name, sales_name in rows:
        review_flags = _resolve_kyc_review_flags(user)
        if not review_flags:
            continue
        if user.blacklist_hit:
            suggested_action = "黑名单拦截"
        elif user.location_risk_blocked:
            suggested_action = "先解除位置风控再处理"
        elif "人脸未完成" in review_flags or "实名未完成" in review_flags:
            suggested_action = "等待人工复核"
        else:
            suggested_action = "待补充资料"
        items.append(
            {
                "id": user.id,
                "phone": user.phone,
                "name": user.name,
                "id_card_num": user.id_card_num,
                "face_auth_status": user.face_auth_status,
                "real_name_status": user.real_name_status,
                "application_submitted_at": user.application_submitted_at,
                "last_login_at": user.last_login_at,
                "source_channel_name": channel_name,
                "source_channel_sales_name": sales_name,
                "review_flags": review_flags,
                "suggested_action": suggested_action,
                "created_at": user.created_at,
            }
        )
    return {
        "total": total,
        "page": skip // limit + 1,
        "size": limit,
        "items": items,
    }


async def _review_kyc_users(
    db: AsyncSession,
    current_admin: Admin,
    user_ids: list[int],
    action: str,
    note: Optional[str] = None,
):
    """批量处理KYC复核结果并记录用户事件。

    :param db: 异步数据库会话
    :param current_admin: 当前管理员
    :param user_ids: 用户ID列表
    :param action: APPROVE或REJECT
    :param note: 审核备注
    :return: 处理数量和用户ID
    """
    ensure_any_admin_page_permission(current_admin, ("users", "kyc-review", "applications"))
    normalized_action = action.upper()
    users = (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all()
    status_value = "VERIFIED" if normalized_action == "APPROVE" else "REJECTED"
    face_value = "APPROVED" if normalized_action == "APPROVE" else "REJECTED"
    title = "KYC人工复核通过" if normalized_action == "APPROVE" else "KYC人工复核拒绝"
    for user in users:
        user.real_name_status = status_value
        user.face_auth_status = face_value
        user.face_auth_at = datetime.now()
        await log_user_event_async(
            db,
            user=user,
            actor_type="ADMIN",
            operator_name=current_admin.username,
            event_type=f"KYC_{normalized_action}",
            title=title,
            detail=(note or "后台KYC复核").strip(),
        )
    await db.commit()
    return {"processed": len(users), "user_ids": [user.id for user in users], "action": normalized_action}


async def _get_monitoring_summary(db: AsyncSession, current_admin: Admin):
    """聚合后台运营与系统监控概览。

    :param db: 异步数据库会话
    :param current_admin: 当前管理员
    :return: 监控汇总
    """
    ensure_any_admin_page_permission(current_admin, ("overview", "financials", "admin-users", "monitoring"))
    today_start, tomorrow = get_today_range()
    jobs = []
    for job in scheduler.get_jobs():
        next_run_time = job.next_run_time
        if next_run_time is not None and next_run_time.tzinfo is not None:
            now_value = datetime.now(next_run_time.tzinfo)
        else:
            now_value = datetime.now()
        jobs.append(
            {
                "job_id": job.id,
                "next_run_time": next_run_time,
                "trigger": str(job.trigger),
                "pending": bool(next_run_time and next_run_time <= now_value),
            }
        )
    return {
        "admin_event_count_24h": int((await db.scalar(select(func.count(UserEvent.id)).where(UserEvent.actor_type == "ADMIN", UserEvent.created_at >= today_start, UserEvent.created_at < tomorrow))) or 0),
        "kyc_pending_count": int((await db.scalar(select(func.count(User.id)).where(or_(User.real_name_status.is_(None), ~func.upper(User.real_name_status).in_(["VERIFIED", "AUTHED", "PASS", "PASSED"]), User.face_auth_status.is_(None), ~func.upper(User.face_auth_status).in_(["APPROVED", "PASS", "PASSED"]), User.location_risk_blocked.is_(True), User.blacklist_hit.is_(True), User.risk_list_hit.is_(True))))) or 0),
        "reminder_event_count_24h": int((await db.scalar(select(func.count(UserEvent.id)).where(UserEvent.event_type == "ADMIN_REMIND", UserEvent.created_at >= today_start, UserEvent.created_at < tomorrow))) or 0),
        "collection_event_count_24h": int((await db.scalar(select(func.count(UserEvent.id)).where(UserEvent.event_type == "ADMIN_COLLECT", UserEvent.created_at >= today_start, UserEvent.created_at < tomorrow))) or 0),
        "momo_pending_count": int((await db.scalar(select(func.count(MomoTransaction.id)).where(MomoTransaction.status == "PENDING"))) or 0),
        "momo_failed_count": int((await db.scalar(select(func.count(MomoTransaction.id)).where(MomoTransaction.status == "FAILED"))) or 0),
        "active_compliance_rule_count": int((await db.scalar(select(func.count(ComplianceRule.id)).where(ComplianceRule.is_active.is_(True)))) or 0),
        "overdue_loan_count": int((await db.scalar(select(func.count(Loan.id)).where(Loan.status == "OVERDUE"))) or 0),
        "scheduled_jobs": jobs,
    }


def _build_message_templates() -> list[dict]:
    return [
        {
            "key": "repay_due_today",
            "title": "到期日提醒",
            "channel": "SMS",
            "trigger": "D0",
            "body": "您的还款今天到期，请尽快完成还款。",
            "enabled": True,
        },
        {
            "key": "repay_overdue_day1",
            "title": "逾期第一天提醒",
            "channel": "SMS",
            "trigger": "D+1",
            "body": "您的账单已逾期 1 天，请尽快处理。",
            "enabled": True,
        },
        {
            "key": "repay_overdue_day3",
            "title": "逾期第三天提醒",
            "channel": "SMS",
            "trigger": "D+3",
            "body": "您的账单已逾期 3 天，请及时联系催收。",
            "enabled": True,
        },
    ]


async def _get_message_center(db: AsyncSession, current_admin: Admin, keyword: Optional[str], skip: int, limit: int):
    """返回消息中心的模板与最近触达记录。

    :param db: 异步数据库会话
    :param current_admin: 当前管理员
    :param keyword: 搜索关键字
    :param skip: 分页偏移量
    :param limit: 每页数量
    :return: 消息中心聚合数据
    """
    ensure_any_admin_page_permission(current_admin, ("overview", "repayments", "collections", "financials", "message-center"))
    limit = min(max(limit, 1), 100)
    templates = _build_message_templates()
    stmt = select(UserEvent).where(UserEvent.actor_type == "ADMIN").where(
        or_(
            UserEvent.event_type == "ADMIN_REMIND",
            UserEvent.event_type == "ADMIN_COLLECT",
            UserEvent.event_type == "ADMIN_COLLECTION_NOTE",
        )
    )
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        stmt = stmt.where(or_(UserEvent.title.like(pattern), UserEvent.detail.like(pattern), UserEvent.operator_name.like(pattern)))
    total = (await db.scalar(select(func.count()).select_from(stmt.subquery()))) or 0
    logs = (
        await db.execute(stmt.order_by(UserEvent.created_at.desc(), UserEvent.id.desc()).offset(skip).limit(limit))
    ).scalars().all()
    today_start, tomorrow = get_today_range()
    reminder_rows = (
        await db.execute(
            select(Loan, User.name, User.phone)
            .join(User, User.id == Loan.user_id)
            .where(
                Loan.status.in_(["DISBURSED", "OVERDUE"]),
                Loan.due_date >= today_start,
                Loan.due_date < tomorrow,
            )
            .order_by(Loan.due_date.asc(), Loan.id.asc())
            .limit(20)
        )
    ).all()
    summary = {
        "template_count": len(templates),
        "enabled_template_count": sum(1 for item in templates if item["enabled"]),
        "recent_message_count": total,
        "reminder_queue_count": len(reminder_rows),
    }
    return {
        "summary": summary,
        "templates": templates,
        "recent_logs": [
            {
                "id": item.id,
                "user_id": item.user_id,
                "loan_id": item.loan_id,
                "actor_type": item.actor_type,
                "title": item.title,
                "detail": item.detail,
                "created_at": item.created_at,
            }
            for item in logs
        ],
        "reminder_queue": [
            {
                "id": loan.id,
                "loan_id": loan.id,
                "user_id": loan.user_id,
                "user_name": user_name,
                "user_phone": user_phone,
                "status": loan.status,
                "due_date": loan.due_date,
                "remaining_repayment_amount": calculate_remaining_repayment_amount(loan),
                "total_repayment_amount": calculate_total_repayment_amount(loan),
            }
            for loan, user_name, user_phone in reminder_rows
        ],
    }

async def _create_admin_user(db: AsyncSession, current_admin: Admin, req: AdminUserCreateRequest):
    ensure_admin_page_permission(current_admin, "admin-users")
    username = req.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="请输入后台用户名")

    if (await db.execute(select(Admin).where(Admin.username == username))).scalar_one_or_none():
        raise HTTPException(status_code=400, detail="后台用户名已存在")

    roles, permissions = resolve_roles_and_permissions(req.roles, req.permissions)
    if not roles:
        raise HTTPException(status_code=400, detail="请至少勾选一个角色")

    admin = Admin(
        username=username,
        password_hash=get_password_hash(req.password),
        roles=serialize_admin_roles(roles),
        permissions=serialize_admin_permissions(permissions),
    )
    db.add(admin)
    await db.commit()
    await db.refresh(admin)
    return serialize_admin_user(admin, current_admin)

async def _update_admin_user(db: AsyncSession, current_admin: Admin, admin_id: int, req: AdminUserUpdateRequest):
    ensure_admin_page_permission(current_admin, "admin-users")
    admin = (await db.execute(select(Admin).where(Admin.id == admin_id))).scalar_one_or_none()
    if not admin:
        raise HTTPException(status_code=404, detail="后台用户不存在")

    payload = req.model_dump(exclude_unset=True)
    if not payload:
        return serialize_admin_user(admin, current_admin)

    if "username" in payload and payload["username"] is not None:
        username = payload["username"].strip()
        if not username:
            raise HTTPException(status_code=400, detail="请输入后台用户名")
        duplicated = (
            await db.execute(select(Admin).where(Admin.username == username, Admin.id != admin_id))
        ).scalar_one_or_none()
        if duplicated:
            raise HTTPException(status_code=400, detail="后台用户名已存在")
        admin.username = username

    if "password" in payload and payload["password"]:
        admin.password_hash = get_password_hash(payload["password"])
    if "is_active" in payload and payload["is_active"] is not None:
        if admin.id == current_admin.id and not payload["is_active"]:
            raise HTTPException(status_code=400, detail="当前登录账号不允许禁用")
        admin.is_active = bool(payload["is_active"])

    if "permissions" in payload:
        roles_input = payload.get("roles", None)
        roles, permissions = resolve_roles_and_permissions(roles_input, payload["permissions"])
        if not roles:
            raise HTTPException(status_code=400, detail="请至少勾选一个角色")
        admin.roles = serialize_admin_roles(roles)
        admin.permissions = serialize_admin_permissions(permissions)
    elif "roles" in payload:
        roles, permissions = resolve_roles_and_permissions(payload["roles"], None)
        if not roles:
            raise HTTPException(status_code=400, detail="请至少勾选一个角色")
        admin.roles = serialize_admin_roles(roles)
        admin.permissions = serialize_admin_permissions(permissions)

    await db.commit()
    await db.refresh(admin)
    return serialize_admin_user(admin, current_admin)

async def _delete_admin_user(db: AsyncSession, current_admin: Admin, admin_id: int):
    ensure_admin_page_permission(current_admin, "admin-users")
    admin = (await db.execute(select(Admin).where(Admin.id == admin_id))).scalar_one_or_none()
    if not admin:
        raise HTTPException(status_code=404, detail="后台用户不存在")
    if admin.id == current_admin.id:
        raise HTTPException(status_code=400, detail="当前登录账号不允许删除")

    await db.delete(admin)
    await db.commit()
    return {"msg": "删除成功"}
